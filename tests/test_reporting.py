"""Exporters, diffing, history store and CLI tests."""

import json
import xml.etree.ElementTree as ET

import pytest

from bom_validator import validate_file
from bom_validator.core.diff import diff_reports
from bom_validator.reporting import exporters
from bom_validator.storage.history import HistoryStore


@pytest.fixture
def report(make_workbook):
    return validate_file(make_workbook())


@pytest.fixture
def broken_report(make_workbook):
    return validate_file(make_workbook(top=[], bot=[]))


class TestExporters:
    @pytest.mark.parametrize("fmt", ["csv", "json", "html", "md", "junit", "xlsx"])
    def test_writes_non_empty_file(self, report, tmp_path, fmt):
        out = exporters.export(report, fmt, tmp_path / f"r.{fmt}")
        assert out.exists()
        assert out.stat().st_size > 200

    def test_json_roundtrip(self, report, tmp_path):
        out = exporters.export(report, "json", tmp_path / "r.json")
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["schema"] == "bom-validation-report/1.0"
        assert len(data["results"]) == report.summary.total_lines
        assert data["summary"]["health_score"] == pytest.approx(
            report.summary.health_score
        )

    def test_junit_is_valid_xml(self, broken_report, tmp_path):
        out = exporters.export(broken_report, "junit", tmp_path / "r.xml")
        root = ET.parse(out).getroot()
        assert root.tag == "testsuites"
        assert int(root.attrib["failures"]) > 0

    def test_csv_has_header_and_rows(self, report, tmp_path):
        out = exporters.export(report, "csv", tmp_path / "r.csv")
        lines = out.read_text(encoding="utf-8-sig").splitlines()
        assert "Stock No" in lines[0]
        assert len(lines) == report.summary.total_lines + 1

    def test_html_contains_data(self, report, tmp_path):
        out = exporters.export(report, "html", tmp_path / "r.html")
        html = out.read_text(encoding="utf-8")
        assert "<table" in html
        assert "1000001" in html
        assert "Health score" in html or "health" in html.lower()

    def test_markdown(self, broken_report, tmp_path):
        out = exporters.export(broken_report, "md", tmp_path / "r.md")
        md = out.read_text(encoding="utf-8")
        assert md.startswith("# BOM Validation Report")
        assert "| Metric | Value |" in md

    def test_excel_sheets(self, report, tmp_path):
        from openpyxl import load_workbook

        out = exporters.export(report, "xlsx", tmp_path / "r.xlsx")
        wb = load_workbook(out)
        assert {"Summary", "Results", "Issues", "Designators"} <= set(wb.sheetnames)
        assert wb["Results"].max_row == report.summary.total_lines + 1

    def test_unknown_format(self, report, tmp_path):
        with pytest.raises(ValueError):
            exporters.export(report, "docx", tmp_path / "r.docx")

    def test_default_filename(self, report):
        name = exporters.default_filename(report, "xlsx")
        assert name.endswith(".xlsx")
        assert "validation" in name
        assert exporters.default_filename(report, "junit").endswith(".xml")


class TestDiff:
    def test_identical_reports(self, make_workbook):
        f = make_workbook()
        d = diff_reports(validate_file(f), validate_file(f))
        assert d.total_changes == 0
        assert d.unchanged == 3
        assert d.health_delta == pytest.approx(0.0)

    def test_added_and_removed(self, make_workbook):
        a = make_workbook("a.xlsx")
        b = make_workbook(
            "b.xlsx",
            lines=[
                ("1", "C1, C2, C3", "Capacitor 100nF", "SMD", "0402", 3, "ACME", "1000001"),
                ("2", "R1, R2", "Resistor 10k", "SMD", "0603", 2, "ACME", "1000002"),
                ("4", "D1", "Diode", "SMD", "SOD", 1, "NXP", "1000009"),
            ],
        )
        d = diff_reports(validate_file(a), validate_file(b))
        assert [x.key for x in d.added] == ["1000009"]
        assert [x.key for x in d.removed] == ["1000003"]

    def test_quantity_change(self, make_workbook):
        a = make_workbook("a.xlsx")
        b = make_workbook(
            "b.xlsx",
            lines=[
                ("1", "C1, C2", "Capacitor 100nF", "SMD", "0402", 2, "ACME", "1000001"),
                ("2", "R1, R2", "Resistor 10k", "SMD", "0603", 2, "ACME", "1000002"),
                ("3", "U1", "MCU STM32", "SMD", "LQFP48", 1, "ST", "1000003"),
            ],
        )
        d = diff_reports(validate_file(a), validate_file(b))
        assert any(c.change == "qty_changed" for c in d.changed)

    def test_markdown_render(self, make_workbook):
        f = make_workbook()
        md = diff_reports(validate_file(f), validate_file(f)).to_markdown()
        assert "# BOM diff" in md


class TestHistory:
    def test_save_and_read(self, report, tmp_path):
        store = HistoryStore(tmp_path / "h.sqlite3")
        run_id = store.save(report, operator="tester")
        assert run_id > 0
        runs = store.recent()
        assert len(runs) == 1
        assert runs[0].operator == "tester"
        assert runs[0].total_lines == report.summary.total_lines

    def test_payload_roundtrip(self, report, tmp_path):
        store = HistoryStore(tmp_path / "h.sqlite3")
        run_id = store.save(report)
        payload = store.payload(run_id)
        assert payload["summary"]["total_lines"] == report.summary.total_lines

    def test_trend(self, report, tmp_path):
        store = HistoryStore(tmp_path / "h.sqlite3")
        for _ in range(4):
            store.save(report)
        from pathlib import Path

        trend = store.trend(Path(report.source_file).name)
        assert len(trend) == 4

    def test_purge_and_delete(self, report, tmp_path):
        store = HistoryStore(tmp_path / "h.sqlite3")
        ids = [store.save(report) for _ in range(6)]
        store.delete(ids[0])
        assert len(store.recent(100)) == 5
        store.purge(keep_last=2)
        assert len(store.recent(100)) == 2

    def test_signoff(self, report, tmp_path):
        store = HistoryStore(tmp_path / "h.sqlite3")
        run_id = store.save(report)
        store.sign_off(run_id, "1000001", "operator-a", "verified visually")
        rows = store.signoffs(run_id)
        assert rows[0]["operator"] == "operator-a"

    def test_stats(self, report, tmp_path):
        store = HistoryStore(tmp_path / "h.sqlite3")
        store.save(report)
        stats = store.stats()
        assert stats["runs"] == 1
        assert stats["lines_checked"] == report.summary.total_lines


class TestCli:
    def test_validate_clean_exits_zero(self, make_workbook, capsys):
        from bom_validator.cli import main

        rc = main(["validate", str(make_workbook()), "--no-history", "--no-color"])
        assert rc == 0
        assert "health" in capsys.readouterr().out

    def test_validate_failing_exits_one(self, make_workbook, capsys):
        from bom_validator.cli import main

        f = make_workbook(top=[], bot=[])
        rc = main(["validate", str(f), "--no-history", "--fail-on", "error"])
        assert rc == 1

    def test_fail_on_never(self, make_workbook):
        from bom_validator.cli import main

        f = make_workbook(top=[], bot=[])
        assert main(["validate", str(f), "--no-history", "--fail-on", "never"]) == 0

    def test_json_only_output(self, make_workbook, capsys):
        from bom_validator.cli import main

        main(["validate", str(make_workbook()), "--no-history", "--json-only"])
        data = json.loads(capsys.readouterr().out)
        assert data["schema"] == "bom-validation-report/1.0"

    def test_report_writing(self, make_workbook, tmp_path, capsys):
        from bom_validator.cli import main

        out = tmp_path / "out.html"
        main([
            "validate", str(make_workbook()), "--no-history",
            "--report", f"html:{out}",
        ])
        assert out.exists()

    def test_inspect(self, make_workbook, capsys):
        from bom_validator.cli import main

        assert main(["inspect", str(make_workbook())]) == 0
        out = capsys.readouterr().out
        assert "header row" in out

    def test_profile_list_and_rules(self, capsys):
        from bom_validator.cli import main

        assert main(["profile", "list"]) == 0
        assert "strict" in capsys.readouterr().out
        assert main(["profile", "rules"]) == 0
        assert "QTY_MISMATCH" in capsys.readouterr().out

    def test_batch(self, make_workbook, tmp_path):
        from bom_validator.cli import main

        d = make_workbook("a.xlsx").parent
        make_workbook("b.xlsx")
        out = tmp_path / "reports"
        rc = main([
            "batch", str(d), "--out", str(out), "--format", "json", "--no-history",
        ])
        assert rc in (0, 1)
        assert (out / "batch_summary.json").exists()
        assert len(list(out.glob("*.json"))) >= 2

    def test_diff_command(self, make_workbook, tmp_path):
        from bom_validator.cli import main

        a, b = make_workbook("a.xlsx"), make_workbook("b.xlsx")
        md = tmp_path / "d.md"
        main(["diff", str(a), str(b), "--md", str(md)])
        assert md.exists()

    def test_missing_file_exits_two(self, tmp_path):
        from bom_validator.cli import main

        assert main(["validate", str(tmp_path / "nope.xlsx"), "--no-history"]) == 2

    def test_version_flag(self):
        from bom_validator.cli import main

        with pytest.raises(SystemExit) as e:
            main(["--version"])
        assert e.value.code == 0
