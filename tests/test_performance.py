"""Correctness tests for the caching / concurrency fast paths.

These do not assert on wall-clock time (that is far too flaky in CI); they
assert that every optimisation produces exactly the same answer as the slow
path it replaces, and that the concurrent paths are safe.
"""

from __future__ import annotations

import threading

import pytest

from bom_validator import ValidationProfile, validate_file
from bom_validator.core import normalize as nz
from bom_validator.core.engine import BomValidationEngine
from bom_validator.io_excel import reader as rd


@pytest.fixture(autouse=True)
def _fresh_caches():
    rd.clear_caches()
    yield
    rd.clear_caches()


class TestWorkbookCache:
    def test_second_load_is_served_from_cache(self, make_workbook):
        f = make_workbook()
        first = rd.WorkbookLoader(f).sheets
        second = rd.WorkbookLoader(f).sheets
        # same objects, not just equal ones
        assert first is second

    def test_cache_can_be_bypassed(self, make_workbook):
        f = make_workbook()
        a = rd.WorkbookLoader(f).sheets
        b = rd.WorkbookLoader(f, use_cache=False).sheets
        assert a is not b
        assert list(a) == list(b)

    def test_modified_file_invalidates_cache(self, make_workbook, tmp_path):
        f = make_workbook()
        before = rd.WorkbookLoader(f).sheet_names()
        from openpyxl import load_workbook

        wb = load_workbook(f)
        wb.create_sheet("extra_sheet")
        wb.save(f)
        after = rd.WorkbookLoader(f).sheet_names()
        assert "extra_sheet" in after
        assert "extra_sheet" not in before

    def test_clear_caches_empties_the_store(self, make_workbook):
        f = make_workbook()
        assert rd.WorkbookLoader(f).sheets
        assert rd.GRID_CACHE.stats()["entries"] == 1
        rd.clear_caches()
        assert rd.GRID_CACHE.stats()["entries"] == 0

    def test_concurrent_first_load_is_consistent(self, make_workbook):
        f = make_workbook()
        results: list[object] = []
        errors: list[BaseException] = []
        barrier = threading.Barrier(6)

        def go() -> None:
            try:
                barrier.wait(timeout=10)
                results.append(rd.WorkbookLoader(f).sheets)
            except BaseException as exc:  # pragma: no cover - failure path
                errors.append(exc)

        threads = [threading.Thread(target=go) for _ in range(6)]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=30)

        assert not errors
        assert len(results) == 6
        assert all(r is results[0] for r in results)


class TestHeaderCache:
    def test_repeated_detection_matches(self, make_workbook):
        f = make_workbook()
        loader = rd.WorkbookLoader(f)
        sheet = loader.sheets["top"]
        p = ValidationProfile()
        a = rd.detect_header(sheet, p, synonyms=p.placement_synonyms)
        b = rd.detect_header(sheet, p, synonyms=p.placement_synonyms)
        assert a.columns == b.columns
        assert a.header_row == b.header_row
        assert a.confidence == b.confidence

    def test_cached_mapping_is_not_shared(self, make_workbook):
        f = make_workbook()
        sheet = rd.WorkbookLoader(f).sheets["top"]
        p = ValidationProfile()
        a = rd.detect_header(sheet, p, synonyms=p.placement_synonyms)
        a.columns["designator"] = 99
        b = rd.detect_header(sheet, p, synonyms=p.placement_synonyms)
        assert b.columns["designator"] != 99

    def test_different_profiles_get_different_mappings(self, make_workbook):
        f = make_workbook()
        sheet = rd.WorkbookLoader(f).sheets["top"]
        base = ValidationProfile()
        forced = ValidationProfile(manual_mapping={"designator": 4})
        auto = rd.detect_header(sheet, base, synonyms=base.placement_synonyms)
        manual = rd.detect_header(sheet, forced, synonyms=forced.placement_synonyms)
        assert auto.columns["designator"] == 0
        assert manual.columns["designator"] == 4


class TestSimilarityGate:
    @pytest.mark.parametrize("threshold", [0.3, 0.5, 0.6, 0.8, 0.88, 0.95])
    def test_gate_never_loses_a_qualifying_pair(self, threshold):
        import random

        random.seed(20240727)
        words = ["cap", "res", "100nf", "10k", "0402", "smd", "x7r", "diode", "a"]
        for _ in range(400):
            a = " ".join(random.choices(words, k=random.randint(1, 5)))
            b = " ".join(random.choices(words, k=random.randint(1, 5)))
            full = nz.similarity(a, b)
            gated = nz.similarity_at_least(a, b, threshold)
            if full >= threshold:
                assert gated == full
            else:
                assert gated in (0.0, full)

    def test_identical_strings(self):
        assert nz.similarity_at_least("abc", "abc", 0.99) == 1.0

    def test_empty_strings(self):
        assert nz.similarity_at_least("", "abc", 0.1) == 0.0
        assert nz.similarity("", "") == 0.0

    def test_similarity_is_memoised(self):
        nz.clear_caches()
        nz.similarity("capacitor 100nf", "capacitor 100pf")
        nz.similarity("capacitor 100nf", "capacitor 100pf")
        assert nz.cache_info()["similarity"][0] >= 1


class TestCleanFastPath:
    @pytest.mark.parametrize(
        "raw",
        [
            None,
            "",
            "plain ascii text",
            "  padded  ",
            "double  space",
            "nan",
            " NaN ",
            "<NA>",
            "1110101.0",
            "C1 - C5",
            "۱۲۳",
            "a\u200bb",
            "a\u00a0b",
            123,
            12.5,
            True,
            "Resistor, 10k, 1%",
        ],
    )
    def test_matches_the_reference_implementation(self, raw):
        import re
        import unicodedata

        def reference(value: object) -> str:
            if value is None:
                return ""
            text = str(value)
            if text.strip().lower() in {"nan", "none", "nat", "<na>"}:
                return ""
            text = unicodedata.normalize("NFKC", text)
            text = nz.strip_zero_width(text)
            text = nz.fold_digits(text)
            text = text.replace("\u00a0", " ")
            return re.sub(r"\s+", " ", text).strip()

        assert nz.clean(raw) == reference(raw)


class TestEngineReuse:
    def test_run_exposes_placements(self, make_workbook):
        f = make_workbook()
        engine = BomValidationEngine()
        report = engine.run(f)
        assert engine.last_placements
        assert len(engine.last_placements) == report.metadata["placements_read"]

    def test_placements_match_a_manual_read(self, make_workbook):
        f = make_workbook()
        engine = BomValidationEngine()
        engine.run(f)
        loader = rd.WorkbookLoader(f)
        p = engine.profile
        buckets = rd.classify_sheets(loader.sheet_names(), p)
        from bom_validator.models import Layer

        manual = []
        for name in buckets["top"]:
            manual += rd.extract_placements(loader.sheets[name], Layer.TOP, p)
        for name in buckets["bot"]:
            manual += rd.extract_placements(loader.sheets[name], Layer.BOT, p)
        assert sorted(x.to_dict().items() for x in engine.last_placements) == sorted(
            x.to_dict().items() for x in manual
        )


class TestDeterminismUnderConcurrency:
    def test_parallel_runs_agree_with_serial(self, make_workbook):
        from concurrent.futures import ThreadPoolExecutor

        files = [make_workbook(f"b{i}.xlsx") for i in range(4)]
        serial = [validate_file(f).to_dict()["results"] for f in files]

        rd.clear_caches()
        with ThreadPoolExecutor(max_workers=4) as ex:
            parallel = [r.to_dict()["results"] for r in ex.map(validate_file, files)]

        assert parallel == serial

    def test_same_file_from_many_threads(self, sample_file):
        from concurrent.futures import ThreadPoolExecutor

        with ThreadPoolExecutor(max_workers=4) as ex:
            reports = list(ex.map(lambda _: validate_file(sample_file), range(4)))
        first = reports[0].to_dict()["results"]
        for r in reports[1:]:
            assert r.to_dict()["results"] == first


class TestParallelBatchCli:
    def _make_folder(self, make_workbook, tmp_path):
        folder = tmp_path / "boards"
        folder.mkdir()
        import shutil

        for i in range(5):
            shutil.copy(make_workbook(f"src{i}.xlsx"), folder / f"board{i}.xlsx")
        return folder

    def test_jobs_resolution(self):
        from bom_validator.cli import _resolve_jobs

        assert _resolve_jobs(1, 10) == 1
        assert _resolve_jobs(3, 10) == 3
        assert _resolve_jobs(9999, 4) == 4
        assert _resolve_jobs(0, 6) >= 1
        assert _resolve_jobs(0, 1) == 1

    @pytest.mark.parametrize("jobs", ["1", "4"])
    def test_batch_runs_with_any_worker_count(
        self, make_workbook, tmp_path, capsys, jobs
    ):
        from bom_validator.cli import main

        folder = self._make_folder(make_workbook, tmp_path)
        out = tmp_path / f"out{jobs}"
        rc = main([
            "batch", str(folder), "-o", str(out), "-f", "json",
            "--no-history", "--no-color", "-j", jobs,
        ])
        assert rc == 0
        assert len(list(out.glob("*.json"))) >= 5
        assert "Processed 5 file(s)" in capsys.readouterr().out

    def test_parallel_and_serial_agree(self, make_workbook, tmp_path):
        import json as _json

        from bom_validator.cli import main

        folder = self._make_folder(make_workbook, tmp_path)
        summaries = {}
        for jobs in ("1", "4"):
            out = tmp_path / f"cmp{jobs}"
            main([
                "batch", str(folder), "-o", str(out), "-f", "json",
                "--no-history", "--no-color", "-j", jobs,
            ])
            summaries[jobs] = _json.loads(
                (out / "batch_summary.json").read_text(encoding="utf-8")
            )
        assert summaries["1"] == summaries["4"]


class TestCacheEviction:
    def test_lru_evicts_oldest(self, make_workbook):
        files = [make_workbook(f"lru{i}.xlsx") for i in range(6)]
        for f in files:
            assert rd.WorkbookLoader(f).sheets
        assert rd.GRID_CACHE.stats()["entries"] == rd.GRID_CACHE.max_entries

    def test_byte_budget_is_respected(self, make_workbook, monkeypatch):
        monkeypatch.setattr(rd.GRID_CACHE, "max_bytes", 1)
        for i in range(3):
            assert rd.WorkbookLoader(make_workbook(f"big{i}.xlsx")).sheets
        # a single entry is always kept so the current file stays hot
        assert rd.GRID_CACHE.stats()["entries"] == 1

    def test_stats_shape(self, make_workbook):
        assert rd.WorkbookLoader(make_workbook()).sheets
        stats = rd.GRID_CACHE.stats()
        assert set(stats) == {"entries", "bytes", "hits", "misses"}
        assert stats["bytes"] > 0
