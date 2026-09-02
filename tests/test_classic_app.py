"""تست‌های واحد برای برنامهٔ کلاسیک تکامل‌یافته (excel scraper.py) و سیستم لایسنس."""

from __future__ import annotations

import hashlib
import hmac
import importlib.util
import json
import os
import sys
import time
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

import bom_classic_core as classic  # noqa: E402  — منطق اصلی برنامه (بدون Qt)
import build_icons  # noqa: E402  — آماده‌سازی آیکون ساخت exe
import license_core  # noqa: E402


def test_gui_file_imports_cleanly():
    """روی سیستم دارای Qt، فایل اصلی برنامه باید بدون خطا import شود."""
    pytest.importorskip("PyQt6.QtGui", exc_type=ImportError,
                        reason="بدون Qt (محیط headless) رد می‌شویم")
    spec = importlib.util.spec_from_file_location(
        "excel_scraper_gui", REPO_ROOT / "excel scraper.py"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    assert hasattr(module, "IndustrialBOMValidator")
    assert hasattr(module, "ActivationDialog")
    assert callable(module.ensure_license)


@pytest.fixture(autouse=True)
def isolated_license_store(tmp_path, monkeypatch):
    """فروشگاه لایسنسِ آزمون را از محیط واقعی جدا می‌کند."""
    data: dict[str, str] = {}
    monkeypatch.setattr(license_core, "_store_set",
                        lambda n, v: data.__setitem__(n, v))
    monkeypatch.setattr(license_core, "_store_get",
                        lambda n: data.get(n))
    monkeypatch.setattr(license_core, "_store_delete",
                        lambda n: data.pop(n, None))
    yield data


# ---------------------------------------------------------------------------
# لایسنس
# ---------------------------------------------------------------------------

class TestLicenseCore:
    DEV = "A" * 40

    def test_roundtrip_ok(self):
        key = license_core.build_license_key(self.DEV, months=3)
        state = license_core.verify_license_key(key, device_id=self.DEV)
        assert state.ok
        assert state.plan == "M3"
        assert 88 <= state.days_left <= 91

    def test_plans(self):
        for months, plan, days in ((1, "M1", 30), (3, "M3", 90), (6, "M6", 180)):
            key = license_core.build_license_key(self.DEV, months)
            state = license_core.verify_license_key(key, device_id=self.DEV)
            assert state.ok and state.plan == plan
            assert days - 1 <= state.days_left <= days

    def test_wrong_device_rejected(self):
        key = license_core.build_license_key(self.DEV, months=1)
        state = license_core.verify_license_key(key, device_id="B" * 40)
        assert not state.ok and state.code == "device"

    def test_tampered_key_rejected(self):
        key = license_core.build_license_key(self.DEV, months=1)
        body = key[5:].replace("-", "")
        idx = 12
        flipped = ("B" if body[idx] == "A" else "A")
        forged = "BOM2-" + body[:idx] + flipped + body[idx + 1:]
        state = license_core.verify_license_key(forged, device_id=self.DEV)
        assert not state.ok
        assert state.code in {"malformed", "signature"}

    def test_expired_rejected(self):
        past = time.time() - 40 * 86400
        key = license_core.build_license_key(self.DEV, months=1, issued_at=past)
        state = license_core.verify_license_key(key, device_id=self.DEV)
        assert not state.ok and state.code == "expired"

    def test_garbage_rejected(self):
        assert license_core.verify_license_key("hello world",
                                               device_id=self.DEV).code == "malformed"
        assert license_core.verify_license_key("",
                                               device_id=self.DEV).code == "malformed"

    def test_activation_cycle(self):
        # بدون لایسنس، برنامه در دورهٔ آزمایشی کار می‌کند
        state = license_core.current_state()
        assert state.ok and state.code == "trial"
        # activate() روی دستگاه جاری عمل می‌کند؛ پس کلید دستگاه جاری می‌سازیم
        real_dev = license_core.get_device_id_raw()
        real_key = license_core.build_license_key(real_dev, months=6)
        ok_state = license_core.activate(real_key)
        assert ok_state.ok
        assert license_core.current_state().ok
        # حذف لایسنس در دورهٔ آزمایشی → بازگشت به حالت آزمایشی
        license_core.deactivate()
        after = license_core.current_state()
        assert after.ok and after.code == "trial"

    def test_clock_tamper_detected(self):
        real_dev = license_core.get_device_id_raw()
        key = license_core.build_license_key(real_dev, months=1)
        assert license_core.activate(key).ok
        now = time.time()
        license_core.note_successful_run(now)
        # دو روز به عقب → نامعتبر
        assert not license_core.check_clock(now - 2 * 86400)
        state = license_core.current_state(now=now - 2 * 86400)
        assert not state.ok and state.code == "clock"
        # دو ساعت اختلاف جزئی → تحمل می‌شود
        assert license_core.check_clock(now - 3600)

    def test_seal_unseal(self):
        blob = license_core._seal_timestamp(1234567890)
        assert license_core._unseal_timestamp(blob) == 1234567890
        assert license_core._unseal_timestamp("1234567891." + blob.split(".")[1]) is None
        assert license_core._unseal_timestamp("garbage") is None

    def test_device_id_format(self):
        dashed = license_core.get_device_id()
        raw = license_core.get_device_id_raw()
        assert raw == dashed.replace("-", "")
        assert len(raw) == 40 and all(c in "0123456789ABCDEF" for c in raw)


class TestTrial:
    """دورهٔ آزمایشی خودکار یک‌ماهه."""

    def test_first_run_starts_trial(self):
        state = license_core.current_state()
        assert state.ok and state.code == "trial"
        assert state.short_label == "نسخهٔ آزمایشی"

    def test_trial_active_for_30_days(self):
        now = time.time()
        license_core._store_set("trial", license_core._seal_trial_stamp(int(now - 29 * 86400)))
        state = license_core.current_state()
        assert state.ok and state.code == "trial"

    def test_trial_expires_after_30_days(self):
        license_core._store_set(
            "trial", license_core._seal_trial_stamp(int(time.time() - 31 * 86400)))
        state = license_core.current_state()
        assert not state.ok and state.code == "trial_expired"
        assert "آزمایشی" in state.reason

    def test_tampered_trial_stamp_blocked(self):
        license_core._store_set("trial", "1234567890.deadbeef")
        state = license_core.current_state()
        assert not state.ok and state.code == "trial_expired"

    def test_license_overrides_expired_trial(self):
        license_core._store_set(
            "trial", license_core._seal_trial_stamp(int(time.time() - 40 * 86400)))
        key = license_core.build_license_key(license_core.get_device_id_raw(), months=1)
        assert license_core.activate(key).ok
        state = license_core.current_state()
        assert state.ok and state.code == "ok"


class TestLicenseUiVisibility:
    """
    قاعدهٔ نمایش لایسنس در پنجرهٔ کاربری: تا وقتی لایسنس وضعیت عادی دارد
    (لایسنس فعال یا دورهٔ آزمایشی) هیچ نشانه‌ای دیده نمی‌شود؛ فقط پس از
    انقضا/خرابی ظاهر و پس از فعال‌سازی موفق دوباره پنهان می‌شود.
    """

    def test_hidden_during_trial(self):
        state = license_core.current_state()
        assert state.code == "trial"
        assert not license_core.license_ui_visible(state)

    def test_hidden_with_active_license(self):
        key = license_core.build_license_key(license_core.get_device_id_raw(), months=3)
        assert license_core.activate(key).ok
        state = license_core.current_state()
        assert state.code == "ok"
        assert not license_core.license_ui_visible(state)

    def test_visible_after_trial_expires(self):
        license_core._store_set(
            "trial", license_core._seal_trial_stamp(int(time.time() - 31 * 86400)))
        state = license_core.current_state()
        assert state.code == "trial_expired"
        assert license_core.license_ui_visible(state)

    def test_visible_on_tampered_data(self):
        license_core._store_set("trial", "1234567890.deadbeef")
        assert license_core.license_ui_visible(license_core.current_state())

    def test_hidden_again_after_activation(self):
        # انقضا → علایم ظاهر می‌شوند؛ فعال‌سازی موفق → دوباره پنهان می‌شوند
        license_core._store_set(
            "trial", license_core._seal_trial_stamp(int(time.time() - 40 * 86400)))
        assert license_core.license_ui_visible(license_core.current_state())
        key = license_core.build_license_key(license_core.get_device_id_raw(), months=6)
        assert license_core.activate(key).ok
        assert not license_core.license_ui_visible(license_core.current_state())


# ---------------------------------------------------------------------------
# منطق Excel کلاسیک
# ---------------------------------------------------------------------------

class TestClassicLogic:
    def _sample_like_df(self) -> pd.DataFrame:
        rows = [
            ["Date:", "04/12/12", None, None, None, None, None, None],
            ["Doc Code :", None, "BOM  (Bill of Material)", None, None, None, None, None],
            ["Item", "Designator", "Part Name", "Part No.", "Type/ Material", "Size", "QTY", "Brand / Supplier "],
            [None, None, None, None, None, None, None, "Stock No."],
            [1, "C100-C103", "Capacitor,X7R,47N,50V", "PN1", "X7R", "0603", 4, 1110101],
            [2, "R1-R2", "Resistor,10K,1%", "PN2", "R", "0603", 2, 1110202],
            [3, "U1", "IC,MCU", "PN3", "IC", "QFP", 1, 1110303],
            ["Item", "Designator", "Part Name", "Part No.", None, None, "QTY", "Stock No."],  # سرتیتر تکراری
            [None, None, None, None, None, None, None, None],
        ]
        return pd.DataFrame(rows)

    def test_find_headers(self):
        df = self._sample_like_df()
        header_row, col_part, col_qty, col_stock = classic.find_first_matching_headers(df)
        assert header_row == 2
        assert col_part == 2
        assert col_qty == 6
        assert col_stock == 7

    def test_extract_data(self):
        df = self._sample_like_df()
        h, cp, cq, cs = classic.find_first_matching_headers(df)
        data = classic.extract_data(df, h, cp, cq, cs)
        assert len(data) == 3  # سرتیتر تکراری و سطر خالی حذف می‌شوند
        assert data[0] == {"Stock": "1110101", "Part Name": "Capacitor,X7R,47N,50V", "Qty": 4}
        assert data[1]["Stock"] == "1110202"
        assert data[2]["Qty"] == 1

    def test_cell_to_str_normalization(self):
        assert classic._cell_to_str(1110101) == "1110101"
        assert classic._cell_to_str(1110101.0) == "1110101"
        assert classic._cell_to_str("1110101.0") == "1110101"
        assert classic._cell_to_str(None) == ""
        assert classic._cell_to_str(float("nan")) == ""
        assert classic._cell_to_str("  C100  ") == "C100"

    def test_flatten_values(self):
        df = pd.DataFrame([
            ["Designator", "SPCO Stock Number"],
            ["C100", 1110101],
            ["C101", 1110101.0],
            [None, 1110202],
        ])
        values = classic._flatten_values(df)
        assert values.count("1110101") == 2
        assert "1110202" in values
        assert "nan" not in values
        assert "C100" in values

    def test_evaluate_rows(self):
        data = [
            {"Stock": "1110101", "Part Name": "Cap", "Qty": 3},
            {"Stock": "1110202", "Part Name": "Res", "Qty": 2},
        ]
        top = ["1110101", "1110101", "1110202"]
        bot = ["1110101"]
        results = classic.evaluate_rows(data, top, bot)
        assert results[0]["Top"] == 2 and results[0]["Bot"] == 1 and results[0]["Valid"]
        assert results[1]["Top"] == 1 and results[1]["Bot"] == 0 and not results[1]["Valid"]


# ---------------------------------------------------------------------------
# خواندن فایل‌های TOP/BOT مجزا + ساخت اکسل خروجی سه‌فایل
# ---------------------------------------------------------------------------

class TestTripleFileMode:
    def _write_xlsx(self, path: Path, sheets: dict[str, list[list]]) -> Path:
        wb = openpyxl.Workbook()
        first = True
        for name, rows in sheets.items():
            ws = wb.active if first else wb.create_sheet(name)
            ws.title = name
            first = False
            for row in rows:
                ws.append(row)
        wb.save(path)
        return path

    @pytest.fixture
    def files(self, tmp_path):
        bom = self._write_xlsx(tmp_path / "bom.xlsx", {
            "مونتاژ ماشینی": [
                ["Date:", "04/12/12", None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                ["Item", "Designator", "Part Name", "Part No.", "Type", "Size", "QTY", "Brand"],
                [None, None, None, None, None, None, None, "Stock No."],
                [1, "C1-C2", "Cap 47N", "PN1", "X7R", "0603", 2, 1110101],
                [2, "R1", "Res 10K", "PN2", "R", "0603", 1, 1110202],
            ],
            "top": [["will", "be", "replaced"]],
            "Notes": [["keep me untouched"]],
        })
        top = self._write_xlsx(tmp_path / "top.xlsx", {
            "top": [
                ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation", "SPCO Stock Number", "Description"],
                ["C1", 10.5, 20.5, 90, 1110101, "Cap 47N"],
                ["C2", 11.5, 21.5, 270, 1110101, "Cap 47N"],
            ]
        })
        bot = self._write_xlsx(tmp_path / "bot.xlsx", {
            "bot": [
                ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation", "SPCO Stock Number", "Description"],
                ["R1", 30.0, 40.0, 180, 1110202, "Res 10K"],
            ]
        })
        return bom, top, bot

    def test_load_placement_values_prefers_named_sheet(self, files, tmp_path):
        _, top, _ = files
        values, sheet = classic.load_placement_values(str(top), "top")
        assert sheet == "top"
        assert values.count("1110101") == 2
        assert "C1" in values

    def test_extract_bom_rows_from_file(self, files):
        bom, _, _ = files
        rows = classic.extract_bom_rows(str(bom))
        assert len(rows) == 2
        assert rows[0]["Stock"] == "1110101"
        assert rows[1]["Qty"] == 1

    def test_build_combined_workbook(self, files, tmp_path):
        bom, top, bot = files
        results = classic.evaluate_rows(
            classic.extract_bom_rows(str(bom)),
            classic.load_placement_values(str(top), "top")[0],
            classic.load_placement_values(str(bot), "bot")[0],
        )
        out = tmp_path / "out.xlsx"
        classic.build_combined_workbook(str(bom), str(top), str(bot), results, str(out))

        wb = openpyxl.load_workbook(out)
        # ۱) ساختار BOM صددرصد حفظ شده (شیت BOM و شیت کناری دست‌نخورده)
        assert "مونتاژ ماشینی" in wb.sheetnames
        assert "Notes" in wb.sheetnames
        assert wb["مونتاژ ماشینی"]["A1"].value == "Date:"
        assert wb["Notes"]["A1"].value == "keep me untouched"
        # ۲) شیت‌های top/bot فقط از فایل‌های جدید
        assert [c.value for c in wb["top"][1]] == [
            "Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation",
            "SPCO Stock Number", "Description",
        ]
        assert wb["top"].max_row == 3
        assert wb["bot"].max_row == 2
        assert wb["top"]["A2"].value == "C1"
        assert wb["top"]["E2"].value == 1110101     # نوع عددی حفظ شده
        assert wb["top"]["B2"].value == 10.5        # مختصات PCB
        # ۳) شیت گزارش اعتبارسنجی
        assert "Validation Report" in wb.sheetnames
        vr = wb["Validation Report"]
        texts = [str(c.value) for row in vr.iter_rows(max_row=3) for c in row if c.value]
        assert any("PASS: 2" in t for t in texts)
        statuses = [vr.cell(row=r, column=6).value for r in range(6, vr.max_row + 1)]
        assert statuses == ["PASS", "PASS"]

    def test_top_bot_values_replaced_not_merged(self, files, tmp_path):
        bom, top, bot = files
        results = classic.evaluate_rows(classic.extract_bom_rows(str(bom)), [], [])
        out = tmp_path / "out2.xlsx"
        classic.build_combined_workbook(str(bom), str(top), str(bot), results, str(out))
        wb = openpyxl.load_workbook(out)
        # محتوای قدیمی شیت top در BOM («will be replaced») دیگر وجود ندارد
        flat = [str(v) for row in wb["top"].iter_rows(values_only=True) for v in row]
        assert "will" not in flat

    def test_combined_locks_g4(self, files, tmp_path):
        bom, top, bot = files
        results = classic.evaluate_rows(
            classic.extract_bom_rows(str(bom)),
            classic.load_placement_values(str(top), "top")[0],
            classic.load_placement_values(str(bot), "bot")[0],
        )
        out = tmp_path / "combined_g4.xlsx"
        classic.build_combined_workbook(str(bom), str(top), str(bot), results, str(out))
        wb = openpyxl.load_workbook(out)
        ws = wb["مونتاژ ماشینی"]
        assert ws["G4"].value == "P.Parsa"
        assert ws["G4"].protection.locked is True
        assert ws["A1"].protection.locked is False
        assert ws.protection.sheet is True
        assert ws.protection.password  # هش رمز ذخیره شده است


class TestThreeOutputs:
    """سه خروجی: BOM مختص هر لایه (کد انبار یکسان + تمام Designatorها) + BOM بازتولیدشده."""

    def _write_xlsx(self, path: Path, sheets: dict[str, list[list]]) -> Path:
        wb = openpyxl.Workbook()
        first = True
        for name, rows in sheets.items():
            ws = wb.active if first else wb.create_sheet(name)
            ws.title = name
            first = False
            for row in rows:
                ws.append(row)
        wb.save(path)
        return path

    @pytest.fixture
    def files(self, tmp_path):
        bom = self._write_xlsx(tmp_path / "main BOM.xlsx", {
            "مونتاژ ماشینی": [
                ["F1 Co", None, None, None, None, "  Provided  by :", "Someone", None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None],
                ["Item", "Designator", "Part Name", "Part No.", "Type", "Size", "QTY", "Brand"],
                [None, None, None, None, None, None, None, "Stock No."],
                [1, "C1-C2", "Cap 47N", "PN1", "X7R", "0603", 3, 1110101],
                [2, "D5", "Cap 47N", "PN1", "X7R", "0603", 1, 1110101],   # کد انبار تکراری → ادغام
                [3, "R1", "Res 10K", "PN2", "R", "0603", 2, 1110202],
                ["یادداشت پایانی", None, None, None, None, None, None, None],
            ],
            "top": [["old", "top"]],
            "bot": [["old", "bot"]],
        })
        top = self._write_xlsx(tmp_path / "top.xlsx", {
            "top": [
                ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation", "SPCO Stock Number", "Description"],
                ["C1", 10.5, 20.5, 90, 1110101, "Cap 47N"],
                ["C2", 11.5, 21.5, 270, 1110101, "Cap 47N"],
                ["C7", 15.0, 25.0, 0, 1110999, "LED RED"],   # کد انبار جدید (در BOM نیست)
            ]
        })
        bot = self._write_xlsx(tmp_path / "bot.xlsx", {
            "bot": [
                ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation", "SPCO Stock Number", "Description"],
                ["R1", 30.0, 40.0, 180, 1110202, "Res 10K"],
                ["R3", 31.0, 41.0, 90, 1110202, "Res 10K"],
            ]
        })
        results = classic.evaluate_rows(
            classic.extract_bom_rows(str(bom)),
            classic.load_placement_values(str(top), "top")[0],
            classic.load_placement_values(str(bot), "bot")[0],
        )
        return bom, top, bot, results

    def _data_rows(self, ws):
        """(stock, qty, designator, part, item, pcb) برای سطرهایی که ستون ۸ عدد دارند."""
        out = []
        for r in range(1, ws.max_row + 1):
            stock = ws.cell(row=r, column=8).value
            if isinstance(stock, (int, float)):
                out.append((int(stock), ws.cell(row=r, column=7).value,
                            ws.cell(row=r, column=2).value,
                            ws.cell(row=r, column=3).value,
                            ws.cell(row=r, column=1).value,
                            ws.cell(row=r, column=9).value))
        return out

    def test_names_are_base_plus_v1(self, files, tmp_path):
        bom, top, bot, results = files
        out_dir = tmp_path / "out"
        out_dir.mkdir()
        paths = classic.build_all_outputs(str(bom), str(top), str(bot), results, str(out_dir))
        assert os.path.basename(paths["top"]) == "main BOM_TOP_v1.xlsx"
        assert os.path.basename(paths["bot"]) == "main BOM_BOT_v1.xlsx"
        assert os.path.basename(paths["bom"]) == "main BOM_v1.xlsx"
        for p in paths.values():
            assert os.path.exists(p)

    def test_top_output_is_layer_bom(self, files):
        bom, top, bot, results = files
        paths = classic.build_all_outputs(
            str(bom), str(top), str(bot), results, os.path.dirname(str(bom)))
        wb = openpyxl.load_workbook(paths["top"])
        assert "bot" not in wb.sheetnames
        assert "top" in wb.sheetnames
        assert wb["top"]["A2"].value == "C1"
        assert wb["top"]["E2"].value == 1110101

        ws = wb["مونتاژ ماشینی"]
        rows = self._data_rows(ws)
        by_stock = {r[0]: r for r in rows}

        # ۱) هر کد انبار دقیقاً یک سطر (کد تکراری ادغام شده)
        assert [r[0] for r in rows].count(1110101) == 1
        # ۲) Designatorهای واقعی لایه + تعداد واقعی لایه (نه تعداد کل BOM)
        assert by_stock[1110101][1] == 2            # QTY = ۲ نه ۳
        assert by_stock[1110101][2] == "C1, C2"
        # ۳) کد انبار جدیدِ موجود فقط در فایل top، سطر تازه می‌گیرد
        assert 1110999 in by_stock
        assert by_stock[1110999][1] == 1
        assert by_stock[1110999][2] == "C7"
        assert by_stock[1110999][3] == "LED RED"    # Part Name از Description لایه
        # شماره‌گذاری Item پشت‌سر و بدون شکاف: ۱ (کد ادغام‌شده) و ۲ (سطر تازه)
        assert by_stock[1110101][4] == 1
        assert by_stock[1110999][4] == 2
        # ۴) قطعهٔ bot-only حذف شده
        assert 1110202 not in by_stock
        # ۵) ستون «pcb» بلافاصله بعد از آخرین عنوان BOM (ستون ۹) و پر از شرح قطعهٔ نقشه
        assert ws.cell(row=8, column=9).value == "pcb"
        assert by_stock[1110101][5] == "Cap 47N"
        assert by_stock[1110999][5] == "LED RED"
        # ۶) سربرگ/عنوان و یادداشت غیرداده حفظ شده
        assert ws["A1"].value == "F1 Co"
        assert any("یادداشت" in str(ws.cell(row=r, column=1).value or "")
                   for r in range(1, ws.max_row + 1))
        assert "Layer Report" in wb.sheetnames

    def test_bot_output_is_layer_bom(self, files):
        bom, top, bot, results = files
        paths = classic.build_all_outputs(
            str(bom), str(top), str(bot), results, os.path.dirname(str(bom)))
        wb = openpyxl.load_workbook(paths["bot"])
        assert "top" not in wb.sheetnames
        assert "bot" in wb.sheetnames
        assert wb["bot"]["A2"].value == "R1"

        ws = wb["مونتاژ ماشینی"]
        rows = self._data_rows(ws)
        by_stock = {r[0]: r for r in rows}
        assert 1110101 not in by_stock
        assert by_stock[1110202][1] == 2
        assert by_stock[1110202][2] == "R1, R3"
        assert ws.cell(row=8, column=9).value == "pcb"
        assert by_stock[1110202][5] == "Res 10K"

    def test_g4_locked_in_all_outputs(self, files):
        bom, top, bot, results = files
        paths = classic.build_all_outputs(
            str(bom), str(top), str(bot), results, os.path.dirname(str(bom)))
        for p in paths.values():
            wb = openpyxl.load_workbook(p)
            ws = wb["مونتاژ ماشینی"]
            assert ws["G4"].value == "P.Parsa", p
            assert ws["G4"].protection.locked is True, p
            assert ws["A1"].protection.locked is False, p
            assert ws.protection.sheet is True, p
            assert ws.protection.password, p


class TestPcbBoardRow:
    """
    سطر «فیبر برد» (Designator=PCB100 / Part Name=«PCB, …») باید در TOP و BOT
    هر دو حفظ شود — حتی اگر جایگاهش در BOM جابجا شده باشد (تشخیص محتوایی).
    """

    def _write_xlsx(self, path: Path, sheets: dict[str, list[list]]) -> Path:
        wb = openpyxl.Workbook()
        first = True
        for name, rows in sheets.items():
            ws = wb.active if first else wb.create_sheet(name)
            ws.title = name
            first = False
            for row in rows:
                ws.append(row)
        wb.save(path)
        return path

    def _make_files(self, tmp_path, pcb_position: str = "last"):
        cap = [1, "C1-C2", "Cap 47N", "PN1", "X7R", "0603", 3, 1110101]
        res = [2, "R1", "Res 10K", "PN2", "R", "0603", 2, 1110202]
        # مثل فایل واقعی، مقادیر سطر فیبر به‌صورت متن ذخیره شده‌اند ('70' / '1')
        pcb = ["70", "PCB100", "PCB, SBMi", "SBMi_F1K_B2",
               "4Layer,FR4", "130 x 150 x 1.6mm", "1", 1100468]
        data = [cap, res, pcb] if pcb_position == "last" else [cap, pcb, res]
        bom = self._write_xlsx(tmp_path / "main BOM.xlsx", {
            "مونتاژ ماشینی": [
                ["F1 Co", None, "Provided by:", None, None, None, None, None],
                [None] * 8, [None] * 8, [None] * 8, [None] * 8, [None] * 8,
                [None] * 8,
                ["Item", "Designator", "Part Name", "Part No.",
                 "Type", "Size", "QTY", "Brand"],
                [None, None, None, None, None, None, None, "Stock No."],
                *data,
                ["یادداشت پایانی", None, None, None, None, None, None, None],
            ],
            "top": [["old", "top"]],
            "bot": [["old", "bot"]],
        })
        top = self._write_xlsx(tmp_path / "top.xlsx", {
            "top": [
                ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation",
                 "SPCO Stock Number", "Description"],
                ["C1", 10.5, 20.5, 90, 1110101, "Cap 47N"],
                ["C2", 11.5, 21.5, 270, 1110101, "Cap 47N"],
                ["C7", 15.0, 25.0, 0, 1110999, "LED RED"],   # کد انبار تازه
            ]
        })
        bot = self._write_xlsx(tmp_path / "bot.xlsx", {
            "bot": [
                ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation",
                 "SPCO Stock Number", "Description"],
                ["R1", 30.0, 40.0, 180, 1110202, "Res 10K"],
                ["R3", 31.0, 41.0, 90, 1110202, "Res 10K"],
            ]
        })
        results = classic.evaluate_rows(
            classic.extract_bom_rows(str(bom)),
            classic.load_placement_values(str(top), "top")[0],
            classic.load_placement_values(str(bot), "bot")[0],
        )
        return bom, top, bot, results

    @staticmethod
    def _rows(ws):
        """(row, stock, qty, designator, part, part_no, item, pcb) سطرهای داده."""
        out = []
        for r in range(1, ws.max_row + 1):
            stock = ws.cell(row=r, column=8).value
            if isinstance(stock, (int, float)):
                out.append((r, int(stock), ws.cell(row=r, column=7).value,
                            ws.cell(row=r, column=2).value,
                            ws.cell(row=r, column=3).value,
                            ws.cell(row=r, column=4).value,
                            ws.cell(row=r, column=1).value,
                            ws.cell(row=r, column=9).value))
        return out

    def test_is_pcb_bom_row_patterns(self):
        f = classic.is_pcb_bom_row
        # موارد مثبت — مستقل از جایگاه ردیف
        assert f("PCB100", "")
        assert f(" pcb 7 ", "")
        assert f("PCB-100", "")
        assert f("", "PCB, SBMi")
        assert f("", "PCB، X")                 # کامای فارسی
        assert f("", "PCB")
        # نشانه‌های جایگزین (وقتی Designator/Name سبک PCB ندارند)
        assert f("", "فیبر برد مدار چاپی")      # کلمهٔ فارسی
        assert f("BOARD", "")
        assert f("P1", "SBMi_F1K_B2", "4Layer,FR4")   # Type ویژهٔ فیبر
        assert f("P2", "main board", "FR4")
        assert f("P3", "panel", "8-layer stackup")
        assert f("", "Bare Board, SBMi")
        assert f("", "anything", "", "SBMi bare-board rev2")
        # موارد منفی — اشتباه گرفته نشوند
        assert not f("C100", "Cap 47N")
        assert not f("", "Socket,PCB Mount")
        assert not f("", "PCBA, assembled")    # PCBA با PCB یکی نیست
        assert not f("", "Res 10K")
        assert not f("", "")
        assert not f("C1", "Cap,MLCC", "Ceramic Multilayer")   # Multilayer ≠ layer فیبر
        assert not f("R1", "Res 10K", "SMD")
        assert not f("J1", "Board-to-Board connector", "SMD")  # کانکتور، نه برد
        assert not f("U1", "IC, MCU", "SMD", "R7F7015834AFP")

    def test_board_row_without_pcb_labels_kept_in_both(self, tmp_path):
        """فیبری که نه PCB در نام دارد و نه Designatorِ PCB — فقط Type=4Layer,FR4."""
        cap = [1, "C1-C2", "Cap 47N", "PN1", "X7R", "0603", 3, 1110101]
        resw = [2, "R1", "Res 10K", "PN2", "R", "0603", 2, 1110202]
        pcb = ["70", "P1", "SBMi_F1K_B2", "SBMi_F1K_B2",
               "4Layer,FR4", "130 x 150 x 1.6mm", "1", 1100468]
        bom = self._write_xlsx(tmp_path / "main BOM.xlsx", {
            "مونتاژ ماشینی": [
                ["F1 Co", None, "Provided by:", None, None, None, None, None],
                [None] * 8, [None] * 8, [None] * 8, [None] * 8, [None] * 8,
                [None] * 8,
                ["Item", "Designator", "Part Name", "Part No.",
                 "Type", "Size", "QTY", "Brand"],
                [None, None, None, None, None, None, None, "Stock No."],
                cap, pcb, resw,      # فیبر حتی وسطِ جدول
                ["یادداشت پایانی", None, None, None, None, None, None, None],
            ],
            "top": [["old", "top"]],
            "bot": [["old", "bot"]],
        })
        top = self._write_xlsx(tmp_path / "top.xlsx", {
            "top": [
                ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation",
                 "SPCO Stock Number", "Description"],
                ["C1", 10.5, 20.5, 90, 1110101, "Cap 47N"],
            ]
        })
        bot = self._write_xlsx(tmp_path / "bot.xlsx", {
            "bot": [
                ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation",
                 "SPCO Stock Number", "Description"],
                ["R1", 30.0, 40.0, 180, 1110202, "Res 10K"],
            ]
        })
        results = classic.evaluate_rows(
            classic.extract_bom_rows(str(bom)),
            classic.load_placement_values(str(top), "top")[0],
            classic.load_placement_values(str(bot), "bot")[0],
        )
        paths = classic.build_all_outputs(
            str(bom), str(top), str(bot), results, str(tmp_path))
        for key in ("top", "bot"):
            wb = openpyxl.load_workbook(paths[key])
            rows = self._rows(wb["مونتاژ ماشینی"])
            by_stock = {r[1]: r for r in rows}
            assert 1100468 in by_stock, f"{key}: فیبر برد گم شده"
            assert by_stock[1100468][3] == "P1"          # Designator دست‌نخورده
            assert [r[1] for r in rows].count(1100468) == 1

    @pytest.mark.parametrize("pcb_position", ["last", "middle"])
    def test_pcb_row_kept_in_both_layers(self, tmp_path, pcb_position):
        bom, top, bot, results = self._make_files(tmp_path, pcb_position)
        out_dir = tmp_path / "out"
        out_dir.mkdir()
        paths = classic.build_all_outputs(
            str(bom), str(top), str(bot), results, str(out_dir))

        for key in ("top", "bot"):
            wb = openpyxl.load_workbook(paths[key])
            ws = wb["مونتاژ ماشینی"]
            rows = self._rows(ws)
            by_stock = {r[1]: r for r in rows}

            # ۱) سطر فیبر برد حذف نشده و دقیقاً یک‌بار آمده
            assert 1100468 in by_stock, f"{key}: فیبر برد گم شده"
            assert [r[1] for r in rows].count(1100468) == 1

            pcb = by_stock[1100468]
            # ۲) مقادیرش عیناً از BOM اصلی مانده (دست‌نخورده)
            assert pcb[3] == "PCB100"                    # Designator
            assert str(pcb[2]).strip() == "1"            # QTY واقعی خود فیبر
            assert pcb[4] == "PCB, SBMi"                 # Part Name
            assert pcb[5] == "SBMi_F1K_B2"               # Part No.
            # ۳) ستون pcb هم شرح فیبر را اعلام می‌کند
            assert pcb[7] == "PCB, SBMi"

            # ۴) شماره‌گذاری Item پشت‌سر و بدون شکاف، شامل سطر فیبر
            items = sorted(r[6] for r in rows)
            assert items == list(range(1, len(rows) + 1)), items

        # ۵) در هر دو لایه قطعاتِ فقط-لایهٔ-مقابل حذف شده‌اند، ولی فیبر مانده
        wb_top = openpyxl.load_workbook(paths["top"])
        top_stocks = {r[1] for r in self._rows(wb_top["مونتاژ ماشینی"])}
        assert 1110202 not in top_stocks and 1100468 in top_stocks
        wb_bot = openpyxl.load_workbook(paths["bot"])
        bot_stocks = {r[1] for r in self._rows(wb_bot["مونتاژ ماشینی"])}
        assert 1110101 not in bot_stocks and 1100468 in bot_stocks

        # ۶) یادداشت غیرداده همچنان حفظ است
        assert any("یادداشت" in str(wb_top["مونتاژ ماشینی"].cell(row=r, column=1).value or "")
                   for r in range(1, wb_top["مونتاژ ماشینی"].max_row + 1))

    def test_pcb_row_stays_below_new_appended_stocks(self, tmp_path):
        """وقتی فیبر آخرین ردیف است، کدهای تازهٔ لایه قبل از آن درج می‌شوند."""
        bom, top, bot, results = self._make_files(tmp_path, "last")
        paths = classic.build_all_outputs(
            str(bom), str(top), str(bot), results, str(tmp_path))
        wb = openpyxl.load_workbook(paths["top"])
        rows = self._rows(wb["مونتاژ ماشینی"])
        by_stock = {r[1]: r for r in rows}
        assert 1110999 in by_stock                  # کد تازهٔ top درج شده
        assert by_stock[1110999][0] < by_stock[1100468][0]   # قبل از سطر فیبر
        assert by_stock[1100468][0] == max(r[0] for r in rows)  # فیبر آخرین داده
        assert by_stock[1100468][6] == len(rows)      # و آخرین شمارهٔ Item

    def test_layer_report_mentions_pcb(self, tmp_path):
        bom, top, bot, results = self._make_files(tmp_path)
        paths = classic.build_all_outputs(
            str(bom), str(top), str(bot), results, str(tmp_path))
        wb = openpyxl.load_workbook(paths["top"])
        report = wb["Layer Report"]
        texts = [str(report.cell(row=r, column=1).value or "")
                 for r in range(1, report.max_row + 1)]
        assert any("فیبر برد" in t and "PCB100" in t for t in texts)


class TestLicenseKeyIntegrity:
    """اطمینان از اینکه payload بدون دسترسی به کلید مخفی قابل جعل نیست."""

    def test_signature_depends_on_master_key(self):
        dev = "C" * 40
        key = license_core.build_license_key(dev, months=1)
        payload_bytes, sig = license_core._decode_key(key)
        wrong = hmac.new(b"attacker-guess", payload_bytes, hashlib.sha256).digest()
        assert not hmac.compare_digest(wrong, sig)
        right = hmac.new(license_core._master_key(), payload_bytes, hashlib.sha256).digest()
        assert hmac.compare_digest(right, sig)

    def test_payload_embeds_device_binding(self):
        dev = "D" * 40
        key = license_core.build_license_key(dev, months=1)
        payload_bytes, _ = license_core._decode_key(key)
        payload = json.loads(payload_bytes.decode())
        assert payload["dev"] == dev
        assert payload["plan"] == "M1"
        assert payload["exp"] - payload["iat"] == 30 * 86400


class TestBuildIcons:
    """آماده‌سازی آیکون برای ساخت exe — یافتن فایلِ نام‌آشنا و تبدیل به ICO."""

    def test_find_icon_missing(self, tmp_path):
        assert build_icons.find_icon_file(str(tmp_path)) is None

    def test_find_users_download_name(self, tmp_path):
        f = tmp_path / "web-data-scraping-icon-svg-download-png-3587064.png"
        f.write_bytes(b"not-really-an-image")
        assert build_icons.find_icon_file(str(tmp_path)) == str(f)

    def test_app_icon_name_preferred(self, tmp_path):
        long_f = tmp_path / "web-data-scraping-icon-svg-download-png-3587064.png"
        short_f = tmp_path / "app_icon.ico"
        long_f.write_bytes(b"a")
        short_f.write_bytes(b"b")
        assert build_icons.find_icon_file(str(tmp_path)) == str(short_f)

    def test_png_to_ico_header(self, tmp_path):
        pytest.importorskip("PIL")
        from PIL import Image
        src = tmp_path / "app_icon.png"
        Image.new("RGBA", (32, 32), (255, 0, 0, 255)).save(src)
        dst = build_icons.png_to_ico(str(src), str(tmp_path / "app_icon.ico"))
        assert Path(dst).read_bytes()[:4] == b"\x00\x00\x01\x00"   # جادوی فایل ICO

    def test_prepare_icons_from_users_png(self, tmp_path):
        pytest.importorskip("PIL")
        from PIL import Image
        src = tmp_path / "web-data-scraping-icon-svg-download-png-3587064.png"
        Image.new("RGBA", (64, 64), (0, 255, 0, 255)).save(src)
        out = tmp_path / "assets"
        icons = build_icons.prepare_icons(str(tmp_path), str(out))
        # آیکون exe به‌صورت ICO تبدیل شده و فایل Runtime با نام ثابت app_icon.png است
        assert Path(icons["exe_icon"]).name == "app_icon.ico"
        assert Path(icons["exe_icon"]).exists()
        assert Path(icons["runtime_icon"]).name == "app_icon.png"
        assert Path(icons["runtime_icon"]).parent == out

    def test_prepare_icons_ico_passthrough(self, tmp_path):
        src = tmp_path / "app_icon.ico"
        src.write_bytes(b"\x00\x00\x01\x00fake-ico-bytes")
        icons = build_icons.prepare_icons(str(tmp_path), str(tmp_path / "assets"))
        # بدون نیاز به تبدیل/بدون Pillow
        assert icons["exe_icon"] == str(src)
        assert Path(icons["runtime_icon"]).name == "app_icon.ico"

    def test_prepare_icons_missing(self, tmp_path):
        icons = build_icons.prepare_icons(str(tmp_path), str(tmp_path / "assets"))
        assert icons == {"exe_icon": None, "runtime_icon": None}
