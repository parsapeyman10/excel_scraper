"""Shared pytest fixtures — synthetic workbooks generated on the fly."""

from __future__ import annotations

import os
from pathlib import Path

import pytest

os.environ.setdefault("BOM_VALIDATOR_HOME", "")

REPO_ROOT = Path(__file__).resolve().parents[1]
SAMPLE = REPO_ROOT / "BOM_TOP_BOT_Component_Deffine.xlsx"


@pytest.fixture(autouse=True)
def isolated_home(tmp_path_factory, monkeypatch):
    """Never touch the developer's real settings/history during tests."""
    home = tmp_path_factory.mktemp("bomhome")
    monkeypatch.setenv("BOM_VALIDATOR_HOME", str(home))
    yield home


@pytest.fixture
def sample_file() -> Path:
    if not SAMPLE.exists():
        pytest.skip("sample workbook not present")
    return SAMPLE


def _write_workbook(path: Path, bom_rows, top_rows, bot_rows, bom_sheet="مونتاژ ماشینی"):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = bom_sheet
    for row in bom_rows:
        ws.append(row)
    for name, rows in (("top", top_rows), ("bot", bot_rows)):
        s = wb.create_sheet(name)
        for row in rows:
            s.append(row)
    wb.save(path)
    return path


PLACEMENT_HEADER = [
    "Designator",
    "Center-X(mm)",
    "Center-Y(mm)",
    "Rotation",
    "SPCO Stock Number",
    "Description",
]

DEFAULT_BOM_LINES = [
    ("1", "C1, C2, C3", "Capacitor 100nF", "SMD", "0402", 3, "ACME", "1000001"),
    ("2", "R1, R2", "Resistor 10k", "SMD", "0603", 2, "ACME", "1000002"),
    ("3", "U1", "MCU STM32", "SMD", "LQFP48", 1, "ST", "1000003"),
]

DEFAULT_TOP_ROWS = [
    ("C1", 10.0, 10.0, 0, "1000001", "Capacitor 100nF"),
    ("C2", 20.0, 10.0, 90, "1000001", "Capacitor 100nF"),
    ("R1", 30.0, 10.0, 0, "1000002", "Resistor 10k"),
    ("U1", 40.0, 40.0, 180, "1000003", "MCU STM32"),
]

DEFAULT_BOT_ROWS = [
    ("C3", 10.0, 50.0, 0, "1000001", "Capacitor 100nF"),
    ("R2", 30.0, 50.0, 270, "1000002", "Resistor 10k"),
]


def _bom_grid(lines, header_offset: int = 2):
    rows = [["Filler banner"] * 3 for _ in range(header_offset)]
    rows.append(
        ["Item", "Designator", "Part Name", "Type/ Material", "Size", "QTY",
         "Brand / Supplier", "Stock No."]
    )
    rows += [list(line) for line in lines]
    return rows


def _save(path: Path, sheets: dict[str, list[list]]):
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = name
        first = False
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


@pytest.fixture
def make_split_workbooks(tmp_path):
    """Factory for the three-file layout: BOM + separate top / bot files."""

    def _make(
        *,
        lines=None,
        top=DEFAULT_TOP_ROWS,
        bot=DEFAULT_BOT_ROWS,
        header_offset: int = 2,
        top_sheet_name: str = "Sheet1",
        bot_sheet_name: str = "Sheet1",
        prefix: str = "",
    ):
        d = tmp_path / (prefix or "split")
        d.mkdir(parents=True, exist_ok=True)
        bom = _save(
            d / "montaj.xlsx",
            {"مونتاژ ماشینی": _bom_grid(lines or DEFAULT_BOM_LINES, header_offset)},
        )
        top_path = bot_path = None
        if top is not None:
            top_path = _save(
                d / "top_export.xlsx",
                {top_sheet_name: [PLACEMENT_HEADER] + [list(r) for r in top]},
            )
        if bot is not None:
            bot_path = _save(
                d / "bot_export.xlsx",
                {bot_sheet_name: [PLACEMENT_HEADER] + [list(r) for r in bot]},
            )
        return bom, top_path, bot_path

    return _make


@pytest.fixture
def make_workbook(tmp_path):
    """Factory returning a minimal but realistic workbook."""

    def _make(
        name: str = "board.xlsx",
        *,
        lines=None,
        top=None,
        bot=None,
        header_offset: int = 2,
    ) -> Path:
        lines = lines if lines is not None else [
            ("1", "C1, C2, C3", "Capacitor 100nF", "SMD", "0402", 3, "ACME", "1000001"),
            ("2", "R1, R2", "Resistor 10k", "SMD", "0603", 2, "ACME", "1000002"),
            ("3", "U1", "MCU STM32", "SMD", "LQFP48", 1, "ST", "1000003"),
        ]
        bom_rows = [["Filler banner"] * 3 for _ in range(header_offset)]
        bom_rows.append(
            ["Item", "Designator", "Part Name", "Type/ Material", "Size", "QTY",
             "Brand / Supplier", "Stock No."]
        )
        bom_rows += [list(line) for line in lines]

        header = ["Designator", "Center-X(mm)", "Center-Y(mm)", "Rotation",
                  "SPCO Stock Number", "Description"]
        default_top = [
            ("C1", 10.0, 10.0, 0, "1000001", "Capacitor 100nF"),
            ("C2", 20.0, 10.0, 90, "1000001", "Capacitor 100nF"),
            ("R1", 30.0, 10.0, 0, "1000002", "Resistor 10k"),
            ("U1", 40.0, 40.0, 180, "1000003", "MCU STM32"),
        ]
        default_bot = [
            ("C3", 10.0, 50.0, 0, "1000001", "Capacitor 100nF"),
            ("R2", 30.0, 50.0, 270, "1000002", "Resistor 10k"),
        ]
        top_rows = [header] + [list(r) for r in (default_top if top is None else top)]
        bot_rows = [header] + [list(r) for r in (default_bot if bot is None else bot)]
        return _write_workbook(tmp_path / name, bom_rows, top_rows, bot_rows)

    return _make
