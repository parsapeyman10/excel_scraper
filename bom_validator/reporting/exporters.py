"""Report exporters: Excel, CSV, JSON, HTML, Markdown, JUnit XML and PDF."""

from __future__ import annotations

import csv
import html
import logging
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape as xml_escape

from ..models import Status, ValidationReport
from ..version import APP_NAME, __version__

log = logging.getLogger(__name__)

RESULT_COLUMNS = [
    ("item", "Item"),
    ("stock_no", "Stock No"),
    ("part_name", "Part Description"),
    ("part_no", "Part No"),
    ("material", "Type/Material"),
    ("size", "Size"),
    ("brand", "Brand/Supplier"),
    ("qty", "Required Qty"),
    ("top_count", "Top Placed"),
    ("bot_count", "Bot Placed"),
    ("placed_total", "Total Placed"),
    ("delta", "Delta"),
    ("status", "Status"),
    ("signed_off", "Signed Off"),
    ("operator_note", "Operator Note"),
    ("source_row", "Source Row"),
]


def _rows(report: ValidationReport) -> Iterable[dict[str, Any]]:
    for r in report.results:
        d = r.to_dict()
        d["issues_text"] = " | ".join(
            f"[{i['severity']}] {i['code']}: {i['message']}" for i in d["issues"]
        )
        d["missing_text"] = ", ".join(d["missing_designators"])
        d["extra_text"] = ", ".join(d["extra_designators"])
        yield d


# ---------------------------------------------------------------------------
# CSV / JSON
# ---------------------------------------------------------------------------


def export_csv(report: ValidationReport, path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    fields = [k for k, _ in RESULT_COLUMNS] + [
        "missing_text",
        "extra_text",
        "issues_text",
    ]
    with p.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow([lbl for _, lbl in RESULT_COLUMNS] + ["Missing", "Extra", "Issues"])
        for row in _rows(report):
            w.writerow([row.get(f, "") for f in fields])
    return p


def export_json(report: ValidationReport, path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(report.to_json(), encoding="utf-8")
    return p


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

_STATUS_FILL = {
    Status.PASS.value: "C6EFCE",
    Status.WARN.value: "FFEB9C",
    Status.FAIL.value: "FFC7CE",
    Status.NOT_PLACED.value: "F8CBAD",
    Status.UNKNOWN.value: "D9D9D9",
}
_STATUS_FONT = {
    Status.PASS.value: "006100",
    Status.WARN.value: "9C6500",
    Status.FAIL.value: "9C0006",
    Status.NOT_PLACED.value: "843C0C",
    Status.UNKNOWN.value: "404040",
}


def export_excel(report: ValidationReport, path: str | Path) -> Path:
    """Multi-sheet, formatted, filterable workbook."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_font = Font(bold=True, color="FFFFFF", size=11)
    head_fill = PatternFill("solid", fgColor="2C3E50")
    title_font = Font(bold=True, size=14, color="1F3864")

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{APP_NAME} — Validation Report"
    ws["A1"].font = title_font
    meta = [
        ("Source file", Path(report.source_file).name),
        ("SHA-256", report.source_sha256),
        ("Profile", report.profile_name),
        ("Generated", report.generated_at.strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Engine version", __version__),
        ("Duration (ms)", round(report.duration_ms, 1)),
        ("BOM sheet", report.mapping.sheet_name),
        ("Header row", report.mapping.header_row + 1),
        ("Mapping confidence", f"{report.mapping.confidence:.0%}"),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)

    s = report.summary
    stats = [
        ("Total BOM lines", s.total_lines),
        ("Passed", s.passed),
        ("Warnings", s.warnings),
        ("Failed", s.failed),
        ("Not placed", s.not_placed),
        ("Required quantity", s.total_required),
        ("Placed quantity", s.total_placed),
        ("Top placements", s.top_placed),
        ("Bottom placements", s.bot_placed),
        ("Orphan placements", s.orphan_placements),
        ("Duplicate designators", s.duplicate_designators),
        ("Pass rate %", round(s.pass_rate, 2)),
        ("Coverage %", round(s.coverage, 2)),
        ("Health score", round(s.health_score, 2)),
    ]
    start = len(meta) + 5
    ws.cell(start - 1, 1, "Statistics").font = title_font
    for i, (k, v) in enumerate(stats, start=start):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70

    # status pie
    pie_top = start + len(stats) + 2
    ws.cell(pie_top, 1, "Status")
    ws.cell(pie_top, 2, "Count")
    for i, (label, value) in enumerate(
        [
            ("PASS", s.passed),
            ("WARN", s.warnings),
            ("FAIL", s.failed),
            ("NOT_PLACED", s.not_placed),
        ],
        start=pie_top + 1,
    ):
        ws.cell(i, 1, label)
        ws.cell(i, 2, value)
    pie = PieChart()
    pie.title = "Verification status distribution"
    pie.add_data(
        Reference(ws, min_col=2, min_row=pie_top, max_row=pie_top + 4), titles_from_data=True
    )
    pie.set_categories(Reference(ws, min_col=1, min_row=pie_top + 1, max_row=pie_top + 4))
    pie.height, pie.width = 8, 14
    ws.add_chart(pie, f"D{pie_top}")

    # ---------------- Results ----------------
    wr = wb.create_sheet("Results")
    headers = [lbl for _, lbl in RESULT_COLUMNS] + ["Missing", "Extra", "Issues"]
    wr.append(headers)
    for c in range(1, len(headers) + 1):
        cell = wr.cell(1, c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    keys = [k for k, _ in RESULT_COLUMNS] + ["missing_text", "extra_text", "issues_text"]
    for row in _rows(report):
        wr.append([row.get(k, "") for k in keys])
        r = wr.max_row
        fill = PatternFill("solid", fgColor=_STATUS_FILL.get(row["status"], "FFFFFF"))
        for c in range(1, len(headers) + 1):
            cell = wr.cell(r, c)
            cell.border = border
            cell.fill = fill
        st = wr.cell(r, keys.index("status") + 1)
        st.font = Font(bold=True, color=_STATUS_FONT.get(row["status"], "000000"))
        st.alignment = Alignment(horizontal="center")
    wr.freeze_panes = "A2"
    wr.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{wr.max_row}"
    widths = [8, 14, 44, 14, 14, 14, 20, 12, 11, 11, 12, 8, 12, 11, 26, 11, 40, 40, 70]
    for i, w in enumerate(widths[: len(headers)], start=1):
        wr.column_dimensions[get_column_letter(i)].width = w

    # bar chart of top offenders
    offenders = sorted(report.failing, key=lambda r: abs(r.delta), reverse=True)[:15]
    if offenders:
        wc = wb.create_sheet("Charts")
        wc.append(["Key", "Delta (placed - required)"])
        for r in offenders:
            wc.append([r.line.key or r.line.part_name[:30], r.delta])
        chart = BarChart()
        chart.title = "Largest quantity deviations"
        chart.y_axis.title = "Delta"
        chart.add_data(
            Reference(wc, min_col=2, min_row=1, max_row=wc.max_row), titles_from_data=True
        )
        chart.set_categories(Reference(wc, min_col=1, min_row=2, max_row=wc.max_row))
        chart.height, chart.width = 10, 24
        wc.add_chart(chart, "D2")
        wc.column_dimensions["A"].width = 34

    # ---------------- Issues ----------------
    wi = wb.create_sheet("Issues")
    wi.append(["Scope", "Severity", "Code", "Key", "Message"])
    for c in range(1, 6):
        cell = wi.cell(1, c)
        cell.font = head_font
        cell.fill = head_fill
    for r in report.results:
        for i in r.issues:
            wi.append(["line", i.severity.label, i.code, i.line_key, i.message])
    for i in report.global_issues:
        wi.append(["global", i.severity.label, i.code, i.line_key, i.message])
    wi.freeze_panes = "A2"
    if wi.max_row > 1:
        wi.auto_filter.ref = f"A1:E{wi.max_row}"
    for col, w in zip("ABCDE", (10, 12, 20, 22, 110), strict=True):
        wi.column_dimensions[col].width = w

    # ---------------- Designator matrix ----------------
    wd = wb.create_sheet("Designators")
    wd.append(["Stock No", "Part", "Layer", "Designator"])
    for c in range(1, 5):
        wd.cell(1, c).font = head_font
        wd.cell(1, c).fill = head_fill
    for r in report.results:
        for d in r.matched_top:
            wd.append([r.line.stock_no, r.line.part_name, "Top", d])
        for d in r.matched_bot:
            wd.append([r.line.stock_no, r.line.part_name, "Bot", d])
    wd.freeze_panes = "A2"
    if wd.max_row > 1:
        wd.auto_filter.ref = f"A1:D{wd.max_row}"
    for col, w in zip("ABCD", (16, 46, 8, 16), strict=True):
        wd.column_dimensions[col].width = w

    # ---------------- Orphans ----------------
    if report.orphan_placements:
        wo = wb.create_sheet("Orphans")
        wo.append(["Designator", "Layer", "Stock No", "Description", "X", "Y", "Rot", "Row"])
        for c in range(1, 9):
            wo.cell(1, c).font = head_font
            wo.cell(1, c).fill = head_fill
        for pl in report.orphan_placements:
            wo.append(
                [
                    pl.designator,
                    pl.layer.label,
                    pl.stock_no,
                    pl.description,
                    pl.x,
                    pl.y,
                    pl.rotation,
                    pl.source_row,
                ]
            )
        wo.freeze_panes = "A2"
        for col, w in zip("ABCDEFGH", (14, 10, 16, 46, 10, 10, 8, 8), strict=True):
            wo.column_dimensions[col].width = w

    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------

_HTML_CSS = """
:root{--pass:#1e8e3e;--warn:#e37400;--fail:#c5221f;--crit:#a50e0e;--ink:#1f2933;
--muted:#6b7280;--line:#e5e7eb;--bg:#f7f8fa;--card:#fff;--accent:#2563eb}
*{box-sizing:border-box}
body{font-family:'Segoe UI',Tahoma,'Iran Sans',system-ui,sans-serif;margin:0;
background:var(--bg);color:var(--ink);font-size:14px}
header{background:linear-gradient(120deg,#1f3864,#2980b9);color:#fff;padding:28px 32px}
header h1{margin:0 0 6px;font-size:24px;letter-spacing:.3px}
header .sub{opacity:.85;font-size:13px}
main{padding:24px 32px;max-width:1600px;margin:0 auto}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:14px;margin-bottom:24px}
.card{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:16px;
box-shadow:0 1px 3px rgba(0,0,0,.05)}
.card .k{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.6px}
.card .v{font-size:28px;font-weight:700;margin-top:6px}
.card.pass .v{color:var(--pass)}.card.warn .v{color:var(--warn)}
.card.fail .v{color:var(--fail)}.card.crit .v{color:var(--crit)}
.bar{height:10px;border-radius:6px;background:var(--line);overflow:hidden;margin-top:10px;display:flex}
.bar span{display:block;height:100%}
h2{font-size:17px;margin:28px 0 12px;border-bottom:2px solid var(--line);padding-bottom:6px}
table{width:100%;border-collapse:collapse;background:var(--card);border:1px solid var(--line);
border-radius:8px;overflow:hidden;font-size:13px}
th{background:#2c3e50;color:#fff;padding:9px 10px;text-align:left;position:sticky;top:0;font-weight:600}
td{padding:7px 10px;border-top:1px solid var(--line);vertical-align:top}
tr:nth-child(even) td{background:#fafbfc}
.status{font-weight:700;padding:2px 9px;border-radius:20px;font-size:11.5px;letter-spacing:.4px}
.s-PASS{background:#e6f4ea;color:var(--pass)}
.s-WARN{background:#fef7e0;color:var(--warn)}
.s-FAIL{background:#fce8e6;color:var(--fail)}
.s-NOT_PLACED{background:#fbe0dc;color:var(--crit)}
.num{text-align:right;font-variant-numeric:tabular-nums}
.muted{color:var(--muted);font-size:12px}
.controls{margin:10px 0;display:flex;gap:10px;flex-wrap:wrap}
input,select{padding:7px 10px;border:1px solid var(--line);border-radius:6px;font-size:13px}
footer{padding:18px 32px;color:var(--muted);font-size:12px;text-align:center}
@media print{header{background:#1f3864!important;-webkit-print-color-adjust:exact}
.controls{display:none}body{font-size:11px}}
"""

_HTML_JS = """
function filterRows(){
 const q=document.getElementById('q').value.toLowerCase();
 const st=document.getElementById('st').value;
 document.querySelectorAll('#results tbody tr').forEach(tr=>{
  const okS=!st||tr.dataset.status===st;
  const okQ=!q||tr.innerText.toLowerCase().includes(q);
  tr.style.display=(okS&&okQ)?'':'none';});
}
function sortBy(idx,numeric){
 const tb=document.querySelector('#results tbody');
 const rows=[...tb.rows];
 const dir=tb.dataset.dir==='asc'?-1:1;tb.dataset.dir=dir===1?'asc':'desc';
 rows.sort((a,b)=>{const x=a.cells[idx].innerText,y=b.cells[idx].innerText;
  return numeric?dir*(parseFloat(x)-parseFloat(y)):dir*x.localeCompare(y);});
 rows.forEach(r=>tb.appendChild(r));
}
"""


def export_html(report: ValidationReport, path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    s = report.summary
    e = html.escape

    def pct(n: int) -> float:
        return (n / s.total_lines * 100) if s.total_lines else 0

    rows = []
    for r in report.results:
        issues = "<br>".join(
            f"<span class='muted'>[{i.severity.label}] {e(i.code)}</span> {e(i.message)}"
            for i in r.issues
        )
        rows.append(
            f"<tr data-status='{r.status.value}'>"
            f"<td>{e(r.line.item)}</td>"
            f"<td><code>{e(r.line.stock_no)}</code></td>"
            f"<td>{e(r.line.part_name)}</td>"
            f"<td class='num'>{r.line.qty}</td>"
            f"<td class='num'>{r.top_count}</td>"
            f"<td class='num'>{r.bot_count}</td>"
            f"<td class='num'>{r.placed_total}</td>"
            f"<td class='num'>{r.delta:+d}</td>"
            f"<td><span class='status s-{r.status.value}'>{r.status.value}</span></td>"
            f"<td>{issues}</td></tr>"
        )

    global_rows = "".join(
        f"<tr><td>{e(i.severity.label)}</td><td><code>{e(i.code)}</code></td>"
        f"<td>{e(i.line_key)}</td><td>{e(i.message)}</td></tr>"
        for i in report.global_issues[:500]
    )

    doc = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>BOM Validation Report — {e(Path(report.source_file).name)}</title>
<style>{_HTML_CSS}</style></head><body>
<header>
 <h1>{e(APP_NAME)} — Validation Report</h1>
 <div class="sub">{e(Path(report.source_file).name)} &nbsp;·&nbsp; profile
 <b>{e(report.profile_name)}</b> &nbsp;·&nbsp;
 {report.generated_at.strftime('%Y-%m-%d %H:%M:%S UTC')} &nbsp;·&nbsp;
 sha256 <code>{e(report.source_sha256[:16])}…</code></div>
</header>
<main>
<div class="cards">
 <div class="card"><div class="k">BOM lines</div><div class="v">{s.total_lines}</div></div>
 <div class="card pass"><div class="k">Passed</div><div class="v">{s.passed}</div></div>
 <div class="card warn"><div class="k">Warnings</div><div class="v">{s.warnings}</div></div>
 <div class="card fail"><div class="k">Failed</div><div class="v">{s.failed}</div></div>
 <div class="card crit"><div class="k">Not placed</div><div class="v">{s.not_placed}</div></div>
 <div class="card"><div class="k">Health score</div><div class="v">{s.health_score:.0f}</div>
  <div class="bar"><span style="width:{s.health_score:.1f}%;background:var(--pass)"></span></div></div>
 <div class="card"><div class="k">Coverage</div><div class="v">{s.coverage:.1f}%</div>
  <div class="muted">{s.total_placed} / {s.total_required} placements</div></div>
 <div class="card"><div class="k">Top / Bottom</div><div class="v">{s.top_placed}/{s.bot_placed}</div></div>
</div>

<div class="bar" title="pass / warn / fail / not placed">
 <span style="width:{pct(s.passed):.2f}%;background:var(--pass)"></span>
 <span style="width:{pct(s.warnings):.2f}%;background:var(--warn)"></span>
 <span style="width:{pct(s.failed):.2f}%;background:var(--fail)"></span>
 <span style="width:{pct(s.not_placed):.2f}%;background:var(--crit)"></span>
</div>

<h2>Line results</h2>
<div class="controls">
 <input id="q" placeholder="Search…" oninput="filterRows()" style="min-width:280px">
 <select id="st" onchange="filterRows()">
  <option value="">All statuses</option><option>PASS</option><option>WARN</option>
  <option>FAIL</option><option value="NOT_PLACED">NOT PLACED</option></select>
</div>
<table id="results"><thead><tr>
 <th onclick="sortBy(0,false)">Item</th><th onclick="sortBy(1,false)">Stock</th>
 <th onclick="sortBy(2,false)">Description</th><th onclick="sortBy(3,true)">Qty</th>
 <th onclick="sortBy(4,true)">Top</th><th onclick="sortBy(5,true)">Bot</th>
 <th onclick="sortBy(6,true)">Placed</th><th onclick="sortBy(7,true)">Δ</th>
 <th onclick="sortBy(8,false)">Status</th><th>Findings</th>
</tr></thead><tbody>{''.join(rows)}</tbody></table>

<h2>Global findings ({len(report.global_issues)})</h2>
<table><thead><tr><th>Severity</th><th>Code</th><th>Key</th><th>Message</th></tr></thead>
<tbody>{global_rows or '<tr><td colspan=4 class=muted>None</td></tr>'}</tbody></table>
</main>
<footer>Generated by {e(APP_NAME)} v{__version__} in {report.duration_ms:.0f} ms ·
sheet <b>{e(report.mapping.sheet_name)}</b>, header row {report.mapping.header_row + 1},
mapping confidence {report.mapping.confidence:.0%}</footer>
<script>{_HTML_JS}</script></body></html>"""
    p.write_text(doc, encoding="utf-8")
    return p


# ---------------------------------------------------------------------------
# Markdown / JUnit / PDF
# ---------------------------------------------------------------------------


def export_markdown(report: ValidationReport, path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    s = report.summary
    lines = [
        f"# BOM Validation Report — {Path(report.source_file).name}",
        "",
        f"*Generated {report.generated_at:%Y-%m-%d %H:%M:%S} UTC · profile "
        f"`{report.profile_name}` · engine v{__version__}*",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "|---|---|",
        f"| BOM lines | {s.total_lines} |",
        f"| Passed | {s.passed} ({s.pass_rate:.1f}%) |",
        f"| Warnings | {s.warnings} |",
        f"| Failed | {s.failed} |",
        f"| Not placed | {s.not_placed} |",
        f"| Required / placed | {s.total_required} / {s.total_placed} ({s.coverage:.1f}%) |",
        f"| Top / bottom | {s.top_placed} / {s.bot_placed} |",
        f"| Orphan placements | {s.orphan_placements} |",
        f"| Health score | **{s.health_score:.1f}/100** |",
        "",
        "## Findings",
        "",
        "| Stock | Description | Qty | Top | Bot | Δ | Status | Issue |",
        "|---|---|--:|--:|--:|--:|---|---|",
    ]
    for r in report.results:
        if r.status is Status.PASS:
            continue
        msg = r.issues[0].message.replace("|", "/") if r.issues else ""
        lines.append(
            f"| `{r.line.stock_no}` | {r.line.part_name[:48]} | {r.line.qty} | "
            f"{r.top_count} | {r.bot_count} | {r.delta:+d} | **{r.status.value}** | {msg} |"
        )
    if report.global_issues:
        lines += ["", "## Global findings", ""]
        for i in report.global_issues[:100]:
            lines.append(f"- **{i.severity.label}** `{i.code}` — {i.message}")
    p.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return p


def export_junit(report: ValidationReport, path: str | Path) -> Path:
    """JUnit XML so CI systems can gate a release on BOM integrity."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    s = report.summary
    cases = []
    for r in report.results:
        name = xml_escape(f"{r.line.stock_no or r.line.part_name}")
        if r.status is Status.PASS:
            cases.append(f'    <testcase classname="bom.line" name="{name}"/>')
        else:
            msg = xml_escape(
                "; ".join(i.message for i in r.issues) or r.status.value
            )
            cases.append(
                f'    <testcase classname="bom.line" name="{name}">\n'
                f'      <failure type="{r.status.value}" message="{msg}"/>\n'
                f"    </testcase>"
            )
    for i in report.global_issues:
        nm = xml_escape(f"{i.code}:{i.line_key}")
        cases.append(
            f'    <testcase classname="bom.global" name="{nm}">\n'
            f'      <failure type="{i.severity.label}" '
            f'message="{xml_escape(i.message)}"/>\n    </testcase>'
        )
    failures = s.failed + s.not_placed + len(report.global_issues)
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        f'<testsuites name="BOM Validation" tests="{len(cases)}" failures="{failures}">\n'
        f'  <testsuite name="{xml_escape(Path(report.source_file).name)}" '
        f'tests="{len(cases)}" failures="{failures}" '
        f'time="{report.duration_ms / 1000:.3f}">\n'
        + "\n".join(cases)
        + "\n  </testsuite>\n</testsuites>\n"
    )
    p.write_text(xml, encoding="utf-8")
    return p


def export_pdf(report: ValidationReport, path: str | Path) -> Path:
    """PDF via Qt's print engine when available, else an HTML sidecar."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp_html = p.with_suffix(".html")
    export_html(report, tmp_html)
    try:
        from PyQt6.QtCore import QMarginsF, QUrl
        from PyQt6.QtGui import QPageLayout, QPageSize, QTextDocument
        from PyQt6.QtPrintSupport import QPrinter

        doc = QTextDocument()
        doc.setHtml(tmp_html.read_text(encoding="utf-8"))
        doc.setBaseUrl(QUrl.fromLocalFile(str(tmp_html.parent) + "/"))
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(str(p))
        printer.setPageLayout(
            QPageLayout(
                QPageSize(QPageSize.PageSizeId.A4),
                QPageLayout.Orientation.Landscape,
                QMarginsF(10, 10, 10, 10),
                QPageLayout.Unit.Millimeter,
            )
        )
        doc.print(printer)
        return p
    except Exception as exc:  # pragma: no cover - depends on Qt runtime
        log.warning("PDF export fell back to HTML: %s", exc)
        return tmp_html


EXPORTERS = {
    "xlsx": export_excel,
    "csv": export_csv,
    "json": export_json,
    "html": export_html,
    "md": export_markdown,
    "junit": export_junit,
    "pdf": export_pdf,
}


def export(report: ValidationReport, fmt: str, path: str | Path) -> Path:
    fn = EXPORTERS.get(fmt.lower().lstrip("."))
    if not fn:
        raise ValueError(f"Unknown export format {fmt!r}. Known: {', '.join(EXPORTERS)}")
    return fn(report, path)


def default_filename(report: ValidationReport, fmt: str) -> str:
    stem = Path(report.source_file).stem or "bom"
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    ext = "xml" if fmt == "junit" else fmt
    return f"{stem}_validation_{ts}.{ext}"
