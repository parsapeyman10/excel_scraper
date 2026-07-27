"""Workbook loading, sheet classification and header auto-detection.

Reads with :mod:`openpyxl` in read-only mode so 100k-row workbooks stay cheap,
and falls back to :mod:`pandas` for legacy ``.xls``/``.csv`` inputs.
"""

from __future__ import annotations

import csv
import logging
import os
import threading
from collections.abc import Callable, Iterator, Sequence
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, replace
from functools import lru_cache
from pathlib import Path
from typing import Any

from ..config import ValidationProfile
from ..core import normalize as nz
from ..models import BomLine, Layer, Placement, SheetMapping

log = logging.getLogger(__name__)

Grid = list[list[Any]]
ProgressCb = Callable[[int, int, str], None] | None


class WorkbookError(RuntimeError):
    """Raised when a workbook cannot be read or understood."""


# ---------------------------------------------------------------------------
# Raw loading
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class SheetData:
    name: str
    rows: Grid
    # memoised header detections keyed by profile fingerprint
    _header_cache: dict[Any, Any] = None  # type: ignore[assignment]
    _width: int = -1

    def __len__(self) -> int:
        return len(self.rows)

    def cell(self, r: int, c: int) -> Any:
        if 0 <= r < len(self.rows):
            row = self.rows[r]
            if 0 <= c < len(row):
                return row[c]
        return None

    @property
    def width(self) -> int:
        if self._width < 0:
            self._width = max((len(r) for r in self.rows), default=0)
        return self._width

    # -- header memoisation -------------------------------------------
    def cached_header(self, key: Any):
        cache = self._header_cache
        return cache.get(key) if cache else None

    def store_header(self, key: Any, mapping: Any) -> None:
        if self._header_cache is None:
            self._header_cache = {}
        self._header_cache[key] = mapping


@lru_cache(maxsize=1)
def _has_lxml() -> bool:
    """openpyxl only releases the GIL while parsing when lxml is installed."""
    try:
        import lxml.etree  # noqa: F401
    except Exception:
        return False
    return True


class _GridCache:
    """Tiny thread-safe LRU of parsed workbooks keyed by (path, mtime, size).

    Opening the same workbook happens several times per run (engine, board
    map, sheet preview). Parsing it once and sharing the immutable grids makes
    the second and third open essentially free.
    """

    def __init__(self, max_entries: int = 4, max_bytes: int = 256 * 1024 * 1024) -> None:
        self._lock = threading.RLock()
        self._data: dict[tuple, dict[str, SheetData]] = {}
        self._order: list[tuple] = []
        self._cost: dict[tuple, int] = {}
        self._loading: dict[tuple, threading.Lock] = {}
        self.max_entries = max_entries
        # keeping whole grids alive is the whole point, but never at the cost
        # of exhausting a shop-floor PC's memory
        self.max_bytes = max_bytes
        self.hits = 0
        self.misses = 0

    def load_lock(self, key: tuple) -> threading.Lock:
        """Serialise concurrent first-time loads of the same workbook.

        The GUI kicks off a preview load and a validation at almost the same
        moment; without this both threads would parse the file.
        """
        with self._lock:
            lock = self._loading.get(key)
            if lock is None:
                lock = self._loading[key] = threading.Lock()
            return lock

    def get(self, key: tuple) -> dict[str, SheetData] | None:
        with self._lock:
            item = self._data.get(key)
            if item is None:
                self.misses += 1
                return None
            self.hits += 1
            self._order.remove(key)
            self._order.append(key)
            return item

    def put(self, key: tuple, value: dict[str, SheetData], cost: int = 0) -> None:
        with self._lock:
            if key in self._data:
                self._order.remove(key)
            self._data[key] = value
            self._cost[key] = cost
            self._order.append(key)
            while self._order and (
                len(self._order) > self.max_entries
                or (
                    len(self._order) > 1
                    and sum(self._cost.values()) > self.max_bytes
                )
            ):
                evicted = self._order.pop(0)
                self._data.pop(evicted, None)
                self._cost.pop(evicted, None)

    def clear(self) -> None:
        with self._lock:
            self._data.clear()
            self._order.clear()
            self._cost.clear()
            self._loading.clear()
            self.hits = self.misses = 0

    def stats(self) -> dict[str, int]:
        with self._lock:
            return {
                "entries": len(self._data),
                "bytes": sum(self._cost.values()),
                "hits": self.hits,
                "misses": self.misses,
            }


GRID_CACHE = _GridCache()


def clear_caches() -> None:
    """Drop every memoised workbook/normalisation result."""
    GRID_CACHE.clear()
    _prepare_synonyms_cached.cache_clear()
    nz.clear_caches()


class WorkbookLoader:
    """Loads any supported tabular file into plain Python grids."""

    EXCEL_EXT = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    LEGACY_EXT = {".xls"}
    TEXT_EXT = {".csv", ".tsv", ".txt"}

    def __init__(
        self,
        path: str | os.PathLike[str],
        max_rows: int = 500_000,
        *,
        use_cache: bool = True,
    ):
        self.path = Path(path)
        self.max_rows = max_rows
        self.use_cache = use_cache
        if not self.path.exists():
            raise WorkbookError(f"File not found: {self.path}")
        self._sheets: dict[str, SheetData] | None = None
        self._lock = threading.RLock()

    # -- public ------------------------------------------------------
    def _cache_key(self) -> tuple:
        st = self.path.stat()
        return (str(self.path.resolve()), st.st_mtime_ns, st.st_size, self.max_rows)

    @property
    def sheets(self) -> dict[str, SheetData]:
        if self._sheets is not None:
            return self._sheets
        with self._lock:
            if self._sheets is not None:  # pragma: no cover - race guard
                return self._sheets
            if not self.use_cache:
                self._sheets = self._load()
                return self._sheets

            key = self._cache_key()
            cached = GRID_CACHE.get(key)
            if cached is not None:
                self._sheets = cached
                return cached

        # parse outside the instance lock, but only one thread per file
        with GRID_CACHE.load_lock(key):
            cached = GRID_CACHE.get(key)
            if cached is None:
                cached = self._load()
                # the source file size is a good, cheap proxy for grid weight
                GRID_CACHE.put(key, cached, cost=key[2] * 40)
        with self._lock:
            self._sheets = cached
        return cached

    def sheet_names(self) -> list[str]:
        return list(self.sheets)

    def get(self, name: str) -> SheetData | None:
        return self.sheets.get(name)

    # -- internals ---------------------------------------------------
    def _load(self) -> dict[str, SheetData]:
        ext = self.path.suffix.lower()
        if ext in self.EXCEL_EXT:
            return self._load_openpyxl()
        if ext in self.LEGACY_EXT:
            return self._load_pandas()
        if ext in self.TEXT_EXT:
            return self._load_text()
        # last resort: let pandas sniff it
        return self._load_pandas()

    # Parallel sheet loading only pays off when the XML parser releases the
    # GIL (lxml) and there are cores to spare — otherwise the extra workbook
    # handles just add overhead, so we stay on the serial path.
    PARALLEL_SHEET_THRESHOLD = 3
    PARALLEL_BYTES_THRESHOLD = 2 * 1024 * 1024
    PARALLEL_MIN_CPUS = 4

    def _read_sheet(self, ws) -> SheetData:
        rows: Grid = []
        append = rows.append
        max_rows = self.max_rows
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                log.warning("Sheet %s truncated at %d rows", ws.title, i)
                break
            # trim trailing empty cells: big win on sheets padded to 16k columns
            end = len(row)
            while end and row[end - 1] is None:
                end -= 1
            append(list(row[:end]) if end != len(row) else list(row))
        return SheetData(ws.title, rows)

    def _load_openpyxl(self) -> dict[str, SheetData]:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise WorkbookError("openpyxl is required to read .xlsx files") from exc
        try:
            wb = openpyxl.load_workbook(
                self.path, data_only=True, read_only=True, keep_links=False
            )
            names = list(wb.sheetnames)
        except Exception as exc:
            raise WorkbookError(f"Cannot open workbook: {exc}") from exc

        try:
            size = self.path.stat().st_size
        except OSError:
            size = 0

        parallel = (
            len(names) >= self.PARALLEL_SHEET_THRESHOLD
            and size >= self.PARALLEL_BYTES_THRESHOLD
            and (os.cpu_count() or 1) >= self.PARALLEL_MIN_CPUS
            and _has_lxml()
        )
        if not parallel:
            try:
                return {ws.title: self._read_sheet(ws) for ws in wb.worksheets}
            finally:
                wb.close()

        wb.close()
        return self._load_openpyxl_parallel(openpyxl, names)

    def _load_openpyxl_parallel(self, openpyxl, names: list[str]) -> dict[str, SheetData]:
        """Read each worksheet from its own read-only workbook handle.

        openpyxl's read-only reader is not thread-safe across a single
        workbook, so every worker opens its own lazy handle. Decompression and
        XML parsing release the GIL often enough that this scales well on
        multi-sheet production workbooks.
        """

        def work(name: str) -> tuple[str, SheetData]:
            wb = openpyxl.load_workbook(
                self.path, data_only=True, read_only=True, keep_links=False
            )
            try:
                return name, self._read_sheet(wb[name])
            finally:
                wb.close()

        workers = min(len(names), (os.cpu_count() or 2), 8)
        out: dict[str, SheetData] = {}
        try:
            with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="sheet") as ex:
                for name, data in ex.map(work, names):
                    out[name] = data
        except Exception as exc:
            log.warning("Parallel sheet load failed (%s); falling back to serial", exc)
            wb = openpyxl.load_workbook(
                self.path, data_only=True, read_only=True, keep_links=False
            )
            try:
                return {ws.title: self._read_sheet(ws) for ws in wb.worksheets}
            finally:
                wb.close()
        # preserve original sheet order
        return {name: out[name] for name in names if name in out}

    def _load_pandas(self) -> dict[str, SheetData]:
        try:
            import pandas as pd
        except ImportError as exc:  # pragma: no cover
            raise WorkbookError("pandas is required for this file type") from exc
        try:
            book = pd.read_excel(self.path, sheet_name=None, header=None)
        except Exception as exc:
            raise WorkbookError(f"Cannot open workbook: {exc}") from exc
        return {
            name: SheetData(name, df.values.tolist()) for name, df in book.items()
        }

    def _load_text(self) -> dict[str, SheetData]:
        delim = "\t" if self.path.suffix.lower() in {".tsv", ".txt"} else ","
        rows: Grid = []
        with self.path.open("r", encoding="utf-8-sig", newline="") as fh:
            for i, row in enumerate(csv.reader(fh, delimiter=delim)):
                if i >= self.max_rows:
                    break
                rows.append(list(row))
        return {self.path.stem: SheetData(self.path.stem, rows)}


# ---------------------------------------------------------------------------
# Sheet classification
# ---------------------------------------------------------------------------


def _matches(name: str, patterns: Sequence[str]) -> bool:
    key = nz.canonical(name)
    return any(nz.canonical(p) in key for p in patterns if p)


def classify_sheets(
    names: Sequence[str], profile: ValidationProfile
) -> dict[str, list[str]]:
    """Bucket sheet names into bom / top / bot / other."""
    buckets: dict[str, list[str]] = {"bom": [], "top": [], "bot": [], "other": []}
    for name in names:
        if _matches(name, profile.bot_sheet_patterns):
            buckets["bot"].append(name)
        elif _matches(name, profile.top_sheet_patterns):
            buckets["top"].append(name)
        elif _matches(name, profile.bom_sheet_patterns):
            buckets["bom"].append(name)
        else:
            buckets["other"].append(name)
    return buckets


def pick_bom_sheet(
    loader: WorkbookLoader, profile: ValidationProfile
) -> str | None:
    """Choose the BOM sheet: name match first, then structural scoring."""
    buckets = classify_sheets(loader.sheet_names(), profile)
    if buckets["bom"]:
        return buckets["bom"][0]
    best, best_score = None, 0.0
    for name in loader.sheet_names():
        if _matches(name, profile.ignore_sheet_patterns):
            continue
        mapping = detect_header(loader.sheets[name], profile)
        if mapping.confidence > best_score:
            best, best_score = name, mapping.confidence
    return best if best_score >= 0.4 else None


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------


def _score_cell(text: str, synonyms: Sequence[str]) -> float:
    """Return 1.0 for exact synonym, 0.75 for containment, else 0."""
    if not text:
        return 0.0
    for s in synonyms:
        if s and text == s:
            return 1.0
    for s in synonyms:
        if len(s) >= 3 and s in text:
            return 0.75
    return 0.0


def _synonym_fingerprint(syns: dict[str, list[str]]) -> tuple:
    """Content-based key so two equal synonym tables share a cache entry."""
    return tuple((k, tuple(v)) for k, v in sorted(syns.items()))


@lru_cache(maxsize=64)
def _prepare_synonyms_cached(fingerprint: tuple) -> list[tuple[str, tuple[str, ...]]]:
    prepared: list[tuple[str, tuple[str, ...]]] = []
    for field_name, options in fingerprint:
        keys = tuple(dict.fromkeys(k for k in (nz.header_key(o) for o in options) if k))
        if keys:
            prepared.append((field_name, keys))
    return prepared


def _prepare_synonyms(syns: dict[str, list[str]]) -> list[tuple[str, tuple[str, ...]]]:
    """Pre-normalise the synonym table once instead of per cell."""
    return _prepare_synonyms_cached(_synonym_fingerprint(syns))


def detect_header(
    sheet: SheetData,
    profile: ValidationProfile,
    synonyms: dict[str, list[str]] | None = None,
) -> SheetMapping:
    """Find the header row and map logical field -> column index.

    Uses a two-pass strategy: score every candidate row, then merge in
    "sub-header" columns that live on the following row(s) — a very common
    pattern in Persian engineering BOMs where ``Stock No.`` sits under a
    merged banner.
    """
    syns = synonyms or profile.column_synonyms
    scan = min(profile.header_scan_rows, len(sheet))
    best = SheetMapping(sheet_name=sheet.name)

    cache_key = (
        _synonym_fingerprint(syns),
        scan,
        profile.header_lookahead_rows,
        tuple(profile.required_columns),
        tuple(sorted(profile.manual_mapping.items())) if profile.manual_mapping else (),
    )
    cached = sheet.cached_header(cache_key)
    if cached is not None:
        # hand out a copy: callers (dialogs, engine) may mutate the mapping
        return replace(cached, columns=dict(cached.columns))

    prepared = _prepare_synonyms(syns)
    row_keys: dict[int, list[tuple[int, str]]] = {}

    def keys_for(rr: int) -> list[tuple[int, str]]:
        """Normalised (column, key) pairs of a row — computed at most once."""
        cachedrow = row_keys.get(rr)
        if cachedrow is None:
            cachedrow = []
            for c, raw in enumerate(sheet.rows[rr]):
                if raw is None:
                    continue
                text = nz.header_key(str(raw))
                if text:
                    cachedrow.append((c, text))
            row_keys[rr] = cachedrow
        return cachedrow

    for r in range(scan):
        found: dict[str, tuple[int, float]] = {}
        for c, text in keys_for(r):
            for field_name, options in prepared:
                score = _score_cell(text, options)
                if score > 0 and (
                    field_name not in found or score > found[field_name][1]
                ):
                    found[field_name] = (c, score)
        if not found:
            continue
        # merge sub-header rows for still-missing fields
        for look in range(1, profile.header_lookahead_rows + 1):
            rr = r + look
            if rr >= len(sheet):
                break
            for c, text in keys_for(rr):
                for field_name, options in prepared:
                    if field_name in found:
                        continue
                    sc = _score_cell(text, options)
                    if sc > 0:
                        found[field_name] = (c, sc * 0.9)

        required = [f for f in profile.required_columns if f in syns]
        hit = sum(1 for f in required if f in found)
        # allow one missing required column, but never accept zero hits
        minimum = max(1, min(len(required), len(required) - 1)) if required else 0
        if required and hit < minimum:
            continue
        confidence = (
            (hit / len(required) if required else 0.0) * 0.7
            + min(len(found) / max(len(syns), 1), 1.0) * 0.3
        )
        if confidence > best.confidence:
            best = SheetMapping(
                sheet_name=sheet.name,
                header_row=r,
                columns={k: v[0] for k, v in found.items()},
                confidence=round(confidence, 4),
                detected_by="auto",
            )
    if profile.manual_mapping:
        best.columns.update(profile.manual_mapping)
        best.detected_by = "manual-override"
    sheet.store_header(cache_key, replace(best, columns=dict(best.columns)))
    return best


# ---------------------------------------------------------------------------
# Row extraction
# ---------------------------------------------------------------------------

_HEADER_ECHO_TOKENS = {
    "stockno",
    "stockid",
    "partname",
    "partdescription",
    "description",
    "qty",
    "quantity",
    "totalrequired",
    "verification",
    "designator",
    "item",
    "partno",
    "brandsupplier",
    "componentname",
    "typematerial",
}


def _is_header_echo(*values: str) -> bool:
    for v in values:
        key = nz.header_key(v)
        if key and key in _HEADER_ECHO_TOKENS:
            return True
    return False


def extract_bom_lines(
    sheet: SheetData,
    mapping: SheetMapping,
    profile: ValidationProfile,
    progress: ProgressCb = None,
) -> tuple[list[BomLine], list[tuple[int, str]]]:
    """Return (lines, skipped) where skipped is (row_index, reason)."""
    if mapping.header_row < 0:
        raise WorkbookError("No header row detected in the BOM sheet")

    col = mapping.columns
    lines: list[BomLine] = []
    skipped: list[tuple[int, str]] = []
    total = len(sheet)

    def cell(r: int, name: str) -> Any:
        idx = col.get(name, -1)
        return sheet.cell(r, idx) if idx >= 0 else None

    for r in range(mapping.header_row + 1, total):
        if progress and r % 250 == 0:
            progress(r, total, "reading BOM")

        part = nz.clean(cell(r, "part_name"))
        stock = nz.clean(cell(r, "stock_no"))
        qty_raw = cell(r, "qty")
        qty_text = nz.clean(qty_raw)
        desig_raw = cell(r, "designator")

        if not any([part, stock, qty_text, nz.clean(desig_raw)]):
            continue
        if _is_header_echo(part, stock, qty_text):
            skipped.append((r, "header echo"))
            continue

        qty = nz.to_int(qty_raw)
        if qty is None:
            # a designator list without qty is still worth keeping
            if desig_raw:
                qty = len(nz.expand_designators(desig_raw))
            else:
                skipped.append((r, f"non-numeric qty {qty_text!r}"))
                continue

        if profile.trim_trailing_dot_zero and stock.endswith(".0"):
            stock = stock[:-2]

        lines.append(
            BomLine(
                item=nz.clean(cell(r, "item")),
                stock_no=stock,
                part_name=part,
                part_no=nz.clean(cell(r, "part_no")),
                material=nz.clean(cell(r, "material")),
                size=nz.clean(cell(r, "size")),
                brand=nz.clean(cell(r, "brand")),
                note=nz.clean(cell(r, "note")),
                qty=qty,
                designators=nz.expand_designators(desig_raw),
                source_row=r + 1,
            )
        )
    if progress:
        progress(total, total, "BOM read complete")
    return lines, skipped


def extract_placements(
    sheet: SheetData,
    layer: Layer,
    profile: ValidationProfile,
    progress: ProgressCb = None,
) -> list[Placement]:
    """Parse a pick-and-place sheet into :class:`Placement` records."""
    mapping = detect_header(
        sheet,
        ValidationProfile(
            header_scan_rows=profile.header_scan_rows,
            header_lookahead_rows=1,
            required_columns=["designator"],
            column_synonyms=profile.placement_synonyms,
        ),
        synonyms=profile.placement_synonyms,
    )
    out: list[Placement] = []
    total = len(sheet)

    if mapping.header_row < 0:
        # Unstructured sheet: treat every non-empty cell as a bare key.
        for r, row in enumerate(sheet.rows):
            for value in row:
                text = nz.clean(value)
                if text:
                    out.append(
                        Placement(
                            designator="",
                            layer=layer,
                            stock_no=text,
                            source_row=r + 1,
                        )
                    )
        return out

    col = mapping.columns
    for r in range(mapping.header_row + 1, total):
        if progress and r % 500 == 0:
            progress(r, total, f"reading {layer.value}")

        def cell(name: str, _row: int = r) -> Any:
            idx = col.get(name, -1)
            return sheet.cell(_row, idx) if idx >= 0 else None

        designator = nz.clean(cell("designator"))
        stock = nz.clean(cell("stock_no"))
        if profile.trim_trailing_dot_zero and stock.endswith(".0"):
            stock = stock[:-2]
        description = nz.clean(cell("description"))
        if not any([designator, stock, description]):
            continue
        if _is_header_echo(designator, stock, description):
            continue

        row_layer = layer
        layer_text = nz.canonical(nz.clean(cell("layer")))
        if layer_text:
            if layer_text.startswith(("b", "bot")):
                row_layer = Layer.BOT
            elif layer_text.startswith(("t", "top")):
                row_layer = Layer.TOP

        out.append(
            Placement(
                designator=designator,
                layer=row_layer,
                stock_no=stock,
                description=description,
                x=nz.to_float(cell("x")),
                y=nz.to_float(cell("y")),
                rotation=nz.to_float(cell("rotation")),
                source_row=r + 1,
            )
        )
    if progress:
        progress(total, total, f"{layer.value} read complete")
    return out


def iter_preview(sheet: SheetData, rows: int = 40, cols: int = 25) -> Iterator[list[Any]]:
    for row in sheet.rows[:rows]:
        yield list(row[:cols])
