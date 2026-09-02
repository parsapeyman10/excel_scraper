"""
منطق مرکزی برنامهٔ کلاسیک BOM Validator (بدون هیچ وابستگی به رابط کاربری)
==========================================================================

این توابع عیناً از «نسخهٔ اولیهٔ 342 خطی» برنامه (کامیت «first items») استخراج
و تکامل یافته‌اند تا هم در رابط گرافیکی («excel scraper.py») و هم در تست‌ها
و محیط‌های بدون GUI قابل استفاده باشند:

* یافتن شیت BOM و ستون‌های Part/Qty/Stock (قواعد فازی نسخهٔ اولیه)
* استخراج ردیف‌های قطعه
* خواندن و تخت‌کردن مقادیر top/bot (از داخل فایل BOM یا فایل‌های مجزا)
* تطبیق و شمارش PASS/FAIL
* ساخت اکسل خروجیِ حالت سه‌فایل: کلون ۱۰۰٪ BOM + مقادیر تازهٔ top/bot با مختصات PCB
"""

from __future__ import annotations

import collections
import contextlib
import datetime
import os
from copy import copy

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Protection
from openpyxl.styles import Font as XlFont

# ثابت‌های خروجی (طبق درخواست کاربر)
G4_FIXED_VALUE = "P.Parsa"        # محتوای همیشگی سلول G4 در همهٔ خروجی‌ها
SHEET_LOCK_PASSWORD = "1373"      # رمز قفل شیت برای عدم تغییر G4
G4_CELL = "G4"

HEADER_KEYWORDS = [
    'stock no', 'stock id', 'stockid', 'part name', 'partname',
    'part description', 'description', 'qty', 'quantity',
    'total required', 'verification', 'ref', 'designator', 'item',
]


# ---------------------------------------------------------------------------
# نرمال‌سازی سلول‌ها
# ---------------------------------------------------------------------------
def _cell_to_str(value) -> str:
    """تبدیل امن مقدار سلول به رشتهٔ تمیز؛ NaN حذف و اعداد صحیح بدون «.0»."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        # اصلاح مصنوع رایج اکسل: «1110101.0» → «1110101»
        head = text[:-2]
        if head.isdigit():
            return head
    return text


def _count_key(text: str) -> str:
    """کلید یکدست برای شمارش؛ هر دو طرفِ تطبیق (BOM و top/bot) با همین قاعده."""
    return text.strip().casefold()


def _flatten_values(df: pd.DataFrame) -> list[str]:
    """تخت‌سازی همهٔ سلول‌های یک شیت به فهرست رشته‌های تمیز (معنای نسخهٔ اولیه)."""
    values: list[str] = []
    for val in df.values.flatten().tolist():
        text = _cell_to_str(val)
        if text:
            values.append(text)
    return values


# ---------------------------------------------------------------------------
# یافتن شیت‌ها و سربرگ‌ها (قواعد اصلی نسخهٔ اولیه)
# ---------------------------------------------------------------------------
def _pick_sheet_name(sheet_names: list[str], needle: str) -> str | None:
    for name in sheet_names:
        if needle in str(name).strip().lower():
            return name
    return None


def find_bom_sheet_name(sheet_names: list[str]) -> str | None:
    """قاعدهٔ نسخهٔ اولیه: شیتی که «مونتاژ» یا «ماشین» در نامش باشد."""
    for sheet in sheet_names:
        if "مونتاژ" in sheet or "ماشین" in sheet:
            return sheet
    return None


def find_first_matching_headers(df: pd.DataFrame):
    """
    یافتن ردیف عنوان و ستون‌های Part/Qty/Stock — منطق اصلی نسخهٔ اولیه.
    خروجی: (header_row, col_part, col_qty, col_stock) — در صورت شکست همه -1.
    """
    header_row = -1
    col_part = -1
    col_qty = -1
    col_stock = -1

    for r_idx in range(min(15, len(df))):
        row = df.iloc[r_idx]
        p_found, q_found = -1, -1

        for c_idx, val in row.items():
            if pd.notna(val):
                val_str = str(val).strip().lower().replace(" ", "").replace("_", "")
                if p_found == -1 and ('partname' in val_str or val_str == 'part'):
                    p_found = c_idx
                elif q_found == -1 and ('qty' in val_str or 'quantity' in val_str):
                    q_found = c_idx

        if p_found != -1 and q_found != -1:
            header_row = r_idx
            col_part = p_found
            col_qty = q_found
            break

    if header_row != -1:
        for check_r in [header_row, header_row + 1]:
            if check_r < len(df):
                row_chk = df.iloc[check_r]
                for c_idx, val in row_chk.items():
                    if pd.notna(val):
                        val_str = str(val).strip().lower().replace(" ", "").replace("_", "")
                        if 'stock' in val_str:
                            col_stock = c_idx
                            break
            if col_stock != -1:
                break

    return header_row, col_part, col_qty, col_stock


def _scan_bom_rows(df: pd.DataFrame, header_row: int, col_part: int,
                   col_qty: int, col_stock: int) -> list[dict]:
    """
    پیمایش همهٔ سطرهای بعد از سربرگ و تشخیص «سطر داده» از «غیر داده» —
    منطق دقیقاً همان extract_data است، به‌علاوهٔ نگاشت شمارهٔ سطر و کلید تطبیق.
    """
    rows: list[dict] = []
    for r_idx in range(header_row + 1, len(df)):
        part_val = df.iat[r_idx, col_part]
        qty_val = df.iat[r_idx, col_qty]
        stock_val = df.iat[r_idx, col_stock]

        is_part_empty = pd.isna(part_val) or str(part_val).strip() == ""
        is_qty_empty = pd.isna(qty_val) or str(qty_val).strip() == ""
        is_stock_empty = pd.isna(stock_val) or str(stock_val).strip() == ""

        if is_part_empty and is_qty_empty and is_stock_empty:
            continue

        part_str = "" if is_part_empty else str(part_val).strip()
        stock_str = "" if is_stock_empty else _cell_to_str(stock_val)
        qty_str = "" if is_qty_empty else _cell_to_str(qty_val)

        # فیلتر پیشرفته برای حذف کامل سرتیترها و مقادیر تکراری عناوین
        stock_lower = stock_str.lower()
        part_lower = part_str.lower()

        is_header_row = any(
            kw in stock_lower or kw in part_lower for kw in HEADER_KEYWORDS
        )

        try:
            qty_num = int(float(qty_str))
        except (ValueError, TypeError):
            is_header_row = True
            qty_num = 0

        rows.append({
            'idx': r_idx,              # اندیس دیتافریم ⇒ سطر ورک‌شیت = idx + 1
            'Stock': stock_str,
            'Part Name': part_str,
            'Qty': qty_num,
            'is_data': not is_header_row,
            'key': _count_key(stock_str if stock_str else part_str),
        })

    return rows


def extract_data(df: pd.DataFrame, header_row: int, col_part: int,
                 col_qty: int, col_stock: int) -> list[dict]:
    """
    استخراج سطرهای قطعه از شیت BOM — منطق اصلی نسخهٔ اولیه
    (به‌علاوهٔ مدیریت امن‌تر سلول‌های عددیِ اکسل).
    """
    return [
        {'Stock': r['Stock'], 'Part Name': r['Part Name'], 'Qty': r['Qty']}
        for r in _scan_bom_rows(df, header_row, col_part, col_qty, col_stock)
        if r['is_data']
    ]


def extract_bom_rows(bom_path: str) -> list[dict]:
    """بازکردن فایل BOM و استخراج ردیف‌ها با همان قواعد نسخهٔ اولیه."""
    xls = pd.ExcelFile(bom_path, engine='openpyxl')
    target_sheet = find_bom_sheet_name(xls.sheet_names)
    if not target_sheet:
        raise ValueError("شیت مرجع 'مونتاژ ماشینی' یافت نشد.")
    df_main = pd.read_excel(xls, sheet_name=target_sheet, header=None)
    header_row, col_part, col_qty, col_stock = find_first_matching_headers(df_main)
    if header_row == -1 or col_part == -1 or col_qty == -1 or col_stock == -1:
        raise ValueError("عناوین اصلی Part Name، Qty و Stock طبق قانون یافت نشدند.")
    return extract_data(df_main, header_row, col_part, col_qty, col_stock)


# ---------------------------------------------------------------------------
# خواندن مقادیر top/bot
# ---------------------------------------------------------------------------
def load_placement_values(path: str, preferred: str) -> tuple[list[str], str]:
    """
    خواندن فایل مجزای TOP/BOT و بازگرداندن فهرست تختِ مقادیر + نام شیت استفاده‌شده.
    اگر شیتی با نام top/bot وجود داشته باشد همان، در غیر این صورت اولین شیت.
    """
    xls = pd.ExcelFile(path, engine='openpyxl')
    sheet = _pick_sheet_name(list(xls.sheet_names), preferred) or xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sheet, header=None)
    return _flatten_values(df), str(sheet)


def load_single_file_values(xls: pd.ExcelFile, sheet_name: str) -> list[str]:
    """معادل اصلاح‌شدهٔ همان خواندن top/bot در نسخهٔ اولیه (از داخل همان فایل BOM)."""
    df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    return _flatten_values(df)


def evaluate_rows(data: list[dict], top_values: list[str],
                  bot_values: list[str]) -> list[dict]:
    """
    محاسبهٔ شمارش TOP/BOT و وضعیت PASS/FAIL برای هر قطعه.
    قاعدهٔ نسخهٔ اولیه: (تعداد TOP + تعداد BOT) == Qty  →  PASS
    """
    top_counter = collections.Counter(_count_key(v) for v in top_values if v)
    bot_counter = collections.Counter(_count_key(v) for v in bot_values if v)

    results = []
    for item in data:
        stock = item['Stock']
        part_name = item['Part Name']
        qty = item['Qty']
        search_key = _count_key(stock if stock else part_name)
        num_top = top_counter.get(search_key, 0) if search_key else 0
        num_bot = bot_counter.get(search_key, 0) if search_key else 0
        is_valid = (num_top + num_bot) == qty
        results.append({
            'Stock': stock, 'Part Name': part_name, 'Qty': qty,
            'Top': num_top, 'Bot': num_bot, 'Status': "PASS" if is_valid else "FAIL",
            'Valid': is_valid,
        })
    return results


# ---------------------------------------------------------------------------
# ساخت اکسل خروجیِ حالت سه‌فایل: کلونِ ۱۰۰٪ BOM + مقادیر تازهٔ top/bot
# ---------------------------------------------------------------------------

_DARK_FILL = PatternFill("solid", start_color="FF34495E", end_color="FF34495E")
_PASS_FILL = PatternFill("solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
_FAIL_FILL = PatternFill("solid", start_color="FFFFC7CE", end_color="FFFFC7CE")
_HEAD_FONT = XlFont(bold=True, color="FFFFFFFF")
_TITLE_FONT = XlFont(bold=True, size=13, color="FF2C3E50")


def _copy_sheet_values(src_ws, dst_ws, header_bold: bool = True) -> None:
    """کپیِ سلول‌به‌سلولِ مقادیر (با حفظ نوع داده) از شیت مبدأ به مقصد."""
    for row in src_ws.iter_rows(values_only=True):
        dst_ws.append(list(row))
    if header_bold and dst_ws.max_row >= 1:
        for cell in dst_ws[1]:
            cell.font = _HEAD_FONT
            cell.fill = _DARK_FILL
            cell.alignment = Alignment(horizontal="center")
    dst_ws.freeze_panes = "A2"
    for col_cells in dst_ws.columns:
        try:
            letter = col_cells[0].column_letter
        except Exception:
            continue
        width = 10
        for cell in col_cells[:200]:
            if cell.value is not None:
                width = max(width, min(42, len(str(cell.value)) + 3))
        dst_ws.column_dimensions[letter].width = width


def build_combined_workbook(bom_path: str, top_path: str, bot_path: str,
                            results: list[dict], out_path: str,
                            top_sheet_name: str = "top",
                            bot_sheet_name: str = "bot") -> str:
    """
    خروجی حالت سه‌فایل: یک کاربرگ جدید که **دقیقاً مثل فایل BOM** است
    (همهٔ شیت‌ها، عنوان‌ها و قالب‌بندی حفظ می‌شود) و فقط:

    * شیت «top» با مقادیر فایل TOP (Designator، مختصات PCB، Rotation، …) جایگذاری می‌شود
    * شیت «bot» با مقادیر فایل BOT جایگذاری می‌شود
    * شیت «Validation Report» با نتیجهٔ PASS/FAIL اضافه می‌شود
    """
    if bom_path.lower().endswith(".xls"):
        raise ValueError("خروجی‌گرفتن فقط برای فایل‌های .xlsx پشتیبانی می‌شود (BOM فعلی .xls است).")

    wb = openpyxl.load_workbook(bom_path)

    for sheet_title, src_path, wanted_index in (
        ("top", top_path, 1),
        ("bot", bot_path, 2),
    ):
        if sheet_title in wb.sheetnames:
            del wb[sheet_title]
        src_wb = openpyxl.load_workbook(src_path, data_only=True)
        src_ws = src_wb[sheet_title] if sheet_title in src_wb.sheetnames else src_wb.active
        index = min(wanted_index, len(wb.sheetnames))
        dst_ws = wb.create_sheet(sheet_title, index)
        _copy_sheet_values(src_ws, dst_ws)

    # ---- شیت گزارش اعتبارسنجی ----
    report_title = "Validation Report"
    if report_title in wb.sheetnames:
        del wb[report_title]
    ws = wb.create_sheet(report_title)
    ws.sheet_view.rightToLeft = True

    now_txt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    passed = sum(1 for r in results if r['Valid'])
    failed = len(results) - passed

    ws.append(["گزارش اعتبارسنجی BOM ↔ TOP/BOT"])
    ws.cell(row=1, column=1).font = _TITLE_FONT
    ws.append([f"تاریخ تولید: {now_txt}",
               f"BOM: {os.path.basename(bom_path)}",
               f"TOP: {os.path.basename(top_path)} (شیت {top_sheet_name})",
               f"BOT: {os.path.basename(bot_path)} (شیت {bot_sheet_name})"])
    ws.append([f"کل قطعات: {len(results)}", f"PASS: {passed}", f"FAIL: {failed}"])
    ws.append([])

    header = ["Stock ID", "Part Name", "Total Required",
              "Top Placements", "Bot Placements", "Verification Status"]
    ws.append(header)
    head_row = ws.max_row
    for cell in ws[head_row]:
        cell.font = _HEAD_FONT
        cell.fill = _DARK_FILL
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        ws.append([r['Stock'], r['Part Name'], r['Qty'], r['Top'], r['Bot'], r['Status']])
        row_idx = ws.max_row
        fill = _PASS_FILL if r['Valid'] else _FAIL_FILL
        font = XlFont(bold=True, color="FF006100" if r['Valid'] else "FF9C0006")
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = fill
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=6).font = font

    widths = [16, 46, 14, 14, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=head_row, column=i).column_letter].width = w
    ws.freeze_panes = f"A{head_row + 1}"

    apply_g4_lock(wb)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# قفل کردن سلول G4 روی مقدار ثابت P.Parsa (با رمز 1373)
# ---------------------------------------------------------------------------
def apply_g4_lock(wb, bom_sheet_name: str | None = None,
                  value: str = G4_FIXED_VALUE,
                  password: str = SHEET_LOCK_PASSWORD) -> None:
    """
    در شیت BOM خروجی:
    * سلول G4 همیشه = «P.Parsa» می‌شود
    * همهٔ سلول‌های دیگر آزاد و فقط G4 قفل می‌شود
    * شیت با رمز محافظت می‌شود تا کسی نتواند G4 را تغییر دهد
    """
    sheet = bom_sheet_name or find_bom_sheet_name(wb.sheetnames)
    if not sheet or sheet not in wb.sheetnames:
        return
    ws = wb[sheet]
    ws[G4_CELL] = value
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            # روی MergedCell خطا می‌دهد — سبک را از سلولِ بالا-چپ ادغام می‌گیرد
            with contextlib.suppress(AttributeError):
                cell.protection = Protection(locked=False)
    ws[G4_CELL].protection = Protection(locked=True)
    ws.protection.sheet = True
    ws.protection.password = password


# ---------------------------------------------------------------------------
# خروجی‌های لایه‌ای: دو اکسل با فرمت BOM که فقط قطعات یک لایه + pcb را دارند
# ---------------------------------------------------------------------------
def parse_placement_records(path: str, preferred: str) -> tuple[list[dict], str, bool]:
    """
    خواندن ساخت‌یافتهٔ فایل TOP/BOT: برای هر سطر
    {designator, stock, description} + نام شیت.

    اگر سربرگِ قابل تشخیص (ستون Designator) پیدا نشود، ok=False برمی‌گردد و
    «بازسازی BOM لایه» انجام نمی‌شود (فقط فیلتر می‌شود).
    """
    xls = pd.ExcelFile(path, engine="openpyxl")
    sheet = _pick_sheet_name(list(xls.sheet_names), preferred) or xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sheet, header=None)

    header_row = -1
    col_des = col_stock = col_desc = -1
    for r_idx in range(min(15, len(df))):
        des_here = -1
        st_here = -1
        dc_here = -1
        for c_idx, val in df.iloc[r_idx].items():
            if pd.isna(val):
                continue
            val_str = str(val).strip().lower().replace(" ", "").replace("_", "")
            if des_here == -1 and 'designator' in val_str:
                des_here = c_idx
            elif st_here == -1 and 'stock' in val_str:
                st_here = c_idx
            elif dc_here == -1 and ('description' in val_str or 'comment' in val_str):
                dc_here = c_idx
        if des_here != -1:
            header_row, col_des, col_stock, col_desc = r_idx, des_here, st_here, dc_here
            break

    if header_row == -1:
        return [], str(sheet), False

    records: list[dict] = []
    for r_idx in range(header_row + 1, len(df)):
        des = _cell_to_str(df.iat[r_idx, col_des])
        stock = _cell_to_str(df.iat[r_idx, col_stock]) if col_stock != -1 else ""
        desc = _cell_to_str(df.iat[r_idx, col_desc]) if col_desc != -1 else ""
        if not des and not stock:
            continue
        if 'designator' in des.lower():
            continue  # سربرگ تکراری
        records.append({
            'designator': des,
            'stock': stock,
            'description': desc,
        })
    return records, str(sheet), True


def aggregate_layer(records: list[dict]) -> list[dict]:
    """
    گروه‌بندی رکوردهای نقشه بر اساس «کد انبار یکسان»:
    خروجی (به ترتیب اولین مشاهده): {key, stock, description, designators[]}
    """
    groups: dict[str, dict] = {}
    for rec in records:
        key = _count_key(rec['stock'] if rec['stock'] else rec['description'])
        if not key:
            continue
        g = groups.setdefault(key, {
            'key': key,
            'stock': rec['stock'],
            'description': rec['description'],
            'designators': [],
        })
        if rec['designator']:
            g['designators'].append(rec['designator'])
    return list(groups.values())


def _find_bom_columns(df: pd.DataFrame):
    """ستون‌های BOM شامل Designator و Item — مکمل find_first_matching_headers."""
    header_row, col_part, col_qty, col_stock = find_first_matching_headers(df)
    col_des = -1
    col_item = -1
    if header_row != -1:
        for c_idx, val in df.iloc[header_row].items():
            if pd.isna(val):
                continue
            val_str = str(val).strip().lower().replace(" ", "").replace("_", "")
            if col_des == -1 and 'designator' in val_str:
                col_des = c_idx
            elif col_item == -1 and val_str == 'item':
                col_item = c_idx
    if col_item == -1:
        col_item = 0
    return header_row, col_part, col_qty, col_stock, col_des, col_item


def _last_bom_title_col(df: pd.DataFrame, header_row: int) -> int:
    """
    آخرین ستونِ دارای عنوان در ناحیهٔ عنوان BOM (از ردیف اول تا زیرسربرگ) — به‌صورت پویا،
    چون در فایل‌های مختلف ممکن است جای متفاوتی باشد ولی معمولاً آخرین ردیف/ستون BOM است.
    خروجی: اندیس ستون (۰-مبنا، قالب دیتافریم).
    """
    last_title = 0
    end = min(header_row + 1, len(df) - 1)
    for r in range(0, end + 1):
        for c_idx, val in df.iloc[r].items():
            if pd.notna(val) and str(val).strip() != "":
                last_title = max(last_title, c_idx)
    return last_title


def _add_pcb_title(ws, df: pd.DataFrame, header_row: int) -> int:
    """عنوان «pcb» را بلافاصله بعد از آخرین عنوان BOM می‌نویسد؛ شمارهٔ ستون ورک‌شیت را برمی‌گرداند."""
    last_title = _last_bom_title_col(df, header_row)
    pcb_col = last_title + 2                      # ورک‌شیت = دیتافریم + ۱، و خود ستون pcb یکی جلوتر
    head_ws_row = header_row + 1
    hcell = ws.cell(row=head_ws_row, column=pcb_col)
    hcell.value = "pcb"
    with contextlib.suppress(Exception):
        hcell._style = copy(ws.cell(row=head_ws_row, column=pcb_col - 1)._style)
    with contextlib.suppress(Exception):
        ws.column_dimensions[hcell.column_letter].width = max(
            34, ws.column_dimensions[hcell.column_letter].width or 0)
    return pcb_col


def _rebuild_bom_sheet_for_layer(ws, df: pd.DataFrame, groups: list[dict]) -> dict:
    """
    بازسازی شیت BOM مخصوص یک لایه:

    * هر کد انبار فقط یک سطر: Designator = همهٔ دیزاینیتورهای همان لایه (با‌هم‌پیوند)
      و QTY = تعداد واقعی آن‌ها در این لایه
    * سطرهای تکراریِ یک کد انبار (در BOM اصلی) ادغام می‌شوند
    * کدهایی که در BOM اصلی نیستند اما در لایه هستند، به انتهای جدول اضافه می‌شوند
    * ستون «pcb» بعد از آخرین عنوان BOM اضافه و شرح قطعهٔ هر کد (از فایل نقشه) در آن نوشته می‌شود
    * سربرگ/عنوان‌ها و سطرهای غیرداده دست‌نخورده می‌مانند
    """
    header_row, col_part, col_qty, col_stock, col_des, col_item = _find_bom_columns(df)
    stats = {"rewritten": 0, "deleted": 0, "appended": 0}
    if header_row == -1 or col_stock == -1 or col_qty == -1:
        return stats

    pcb_col = _add_pcb_title(ws, df, header_row)   # ستون «pcb» — بعد از آخرین عنوان BOM

    by_key = {g['key']: g for g in groups}
    scans = _scan_bom_rows(df, header_row, col_part, col_qty, col_stock)

    seen: set[str] = set()
    delete_ws_rows: list[int] = []
    data_rows = [s for s in scans if s['is_data']]
    for s in data_rows:
        ws_r = s['idx'] + 1
        if s['key'] not in by_key or s['key'] in seen:
            delete_ws_rows.append(ws_r)
            continue
        seen.add(s['key'])
        g = by_key[s['key']]
        ws.cell(row=ws_r, column=col_qty + 1).value = len(g['designators'])
        if col_des != -1:
            ws.cell(row=ws_r, column=col_des + 1).value = ", ".join(g['designators'])
        ws.cell(row=ws_r, column=pcb_col).value = g['description'] or s['Part Name']
        stats["rewritten"] += 1

    # افزودن کدهای انبارِ موجود در لایه ولی غایب از BOM اصلی
    new_groups = [g for g in groups if g['key'] not in seen]
    if new_groups and data_rows:
        ref_row = data_rows[-1]['idx'] + 1          # آخرین سطر دادهٔ اصلی (قبل از حذف‌ها)
        ws.insert_rows(ref_row + 1, amount=len(new_groups))
        style_cols = {col_item, col_des, col_part, col_qty, col_stock, pcb_col - 1}
        for offset, g in enumerate(new_groups):
            row = ref_row + 1 + offset
            for c in style_cols:
                if c == -1:
                    continue
                with contextlib.suppress(Exception):
                    ws.cell(row=row, column=c + 1)._style = copy(ws.cell(row=ref_row, column=c + 1)._style)
            if col_des != -1:
                ws.cell(row=row, column=col_des + 1).value = ", ".join(g['designators'])
            if col_part != -1:
                ws.cell(row=row, column=col_part + 1).value = g['description']
            ws.cell(row=row, column=col_qty + 1).value = len(g['designators'])
            stock_val = g['stock']
            if isinstance(stock_val, str) and stock_val.isdigit():
                stock_val = int(stock_val)   # هم‌نوع با سلول‌های کد انبار در BOM
            ws.cell(row=row, column=col_stock + 1).value = stock_val
            ws.cell(row=row, column=pcb_col).value = g['description']
            stats["appended"] += 1

    # حذف سطرها از پایین به بالا (شماره‌ها مبتنی بر ساختار اصلی، زیر نقطهٔ درج)
    for ws_r in sorted(set(delete_ws_rows), reverse=True):
        ws.delete_rows(ws_r, 1)
        stats["deleted"] += 1

    # شماره‌گذاری مجددِ پشت‌سرِ ستون Item (بدون شکافِ ناشی از حذف/درج)
    _renumber_items(ws, header_row, col_item, col_stock, col_qty)
    return stats


def _renumber_items(ws, header_row0: int, col_item0: int,
                    col_stock0: int, col_qty0: int) -> None:
    """شماره‌گذاری متوالی ستون Item فقط برای سطرهای داده (سربرگ/فوتر/خالی رد می‌شوند)."""
    if header_row0 < 0 or col_item0 < 0:
        return
    n = 0
    start_row = header_row0 + 2  # ws: header_row0+1 است خود سربرگ؛ از زیر آن اسکن کن
    for r in range(start_row, ws.max_row + 1):
        stock = ws.cell(row=r, column=col_stock0 + 1).value
        qty = ws.cell(row=r, column=col_qty0 + 1).value
        if stock in (None, "") or not isinstance(qty, (int, float)):
            continue
        if any(kw in str(stock).strip().lower() for kw in HEADER_KEYWORDS):
            continue  # زیرسربرگ «Stock No.»
        n += 1
        ws.cell(row=r, column=col_item0 + 1).value = n


def _add_layer_report_sheet(wb, layer: str, groups: list[dict],
                            bom_keys: set[str], placement_path: str) -> None:
    """گزارش لایه: کد انبار، نام قطعه، تعداد دیزاینیتور و وضعیت حضور در BOM اصلی."""
    title = "Layer Report"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True

    now_txt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    in_bom = sum(1 for g in groups if g['key'] in bom_keys)

    ws.append([f"گزارش لایهٔ {layer.upper()} — کدهای انبار + تمام Designatorهای لایه"])
    ws.cell(row=1, column=1).font = _TITLE_FONT
    ws.append([f"تاریخ تولید: {now_txt}",
               f"{layer.upper()}: {os.path.basename(placement_path)}"])
    ws.append([f"کدهای انبار لایه: {len(groups)}",
               f"موجود در BOM اصلی: {in_bom}",
               f"جدید (فقط در فایل لایه): {len(groups) - in_bom}"])
    ws.append([])

    header = ["Stock ID", "Part Name", "Designator Count", "Designators", "In Original BOM"]
    ws.append(header)
    head_row = ws.max_row
    for cell in ws[head_row]:
        cell.font = _HEAD_FONT
        cell.fill = _DARK_FILL
        cell.alignment = Alignment(horizontal="center")
    for g in groups:
        present = g['key'] in bom_keys
        ws.append([g['stock'], g['description'], len(g['designators']),
                   ", ".join(g['designators']), "بله" if present else "خیر"])
        row_idx = ws.max_row
        fill = _PASS_FILL if present else _FAIL_FILL
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = fill
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="left")
    widths = [16, 40, 16, 60, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=head_row, column=i).column_letter].width = w
    ws.freeze_panes = f"A{head_row + 1}"


def build_layer_workbook(bom_path: str, placement_path: str, layer: str,
                         out_path: str) -> str:
    """
    ساخت اکسلِ یک لایه با فرمت کامل BOM — «BOM مختص آن لایه»:

    * همهٔ ساختار/عنوان فایل BOM حفظ می‌شود
    * شیت BOM بر اساس «کد انبار یکسان» بازسازی می‌شود: برای هر کد، همهٔ
      Designatorهای واقعی آن لایه + تعدادشان (از داخل فایل TOP/BOT)
    * شیت لایهٔ مقابل حذف و شیت همین لایه با مقادیر فایل (Designator + مختصات PCB) پر می‌شود
    * شیت «Layer Report» اضافه و سلول G4 قفل می‌گردد
    """
    if layer not in ("top", "bot"):
        raise ValueError("layer باید 'top' یا 'bot' باشد")
    if bom_path.lower().endswith(".xls"):
        raise ValueError("خروجی‌گرفتن فقط برای فایل‌های .xlsx پشتیبانی می‌شود.")

    other = "bot" if layer == "top" else "top"
    wb = openpyxl.load_workbook(bom_path)

    # ۱) حذف شیت لایهٔ مقابل (فقط قطعات همین لایه در خروجی باشند)
    for name in list(wb.sheetnames):
        if other in str(name).strip().lower():
            del wb[name]

    # ۲) قرار دادن مقادیر تازهٔ این لایه (با مختصات PCB)
    if layer in wb.sheetnames:
        del wb[layer]
    src_wb = openpyxl.load_workbook(placement_path, data_only=True)
    src_ws = src_wb[layer] if layer in src_wb.sheetnames else src_wb.active
    dst_ws = wb.create_sheet(layer, min(1, len(wb.sheetnames)))
    _copy_sheet_values(src_ws, dst_ws)

    # ۳) بازسازی شیت BOM بر اساس کدهای انبار این لایه
    bom_sheet = find_bom_sheet_name(wb.sheetnames)
    records, _, records_ok = parse_placement_records(placement_path, layer)
    groups = aggregate_layer(records) if records_ok else []
    bom_keys: set[str] = set()
    if bom_sheet:
        df = pd.read_excel(bom_path, sheet_name=bom_sheet, header=None, engine="openpyxl")
        header_row, col_part, col_qty, col_stock = find_first_matching_headers(df)
        if header_row != -1:
            bom_keys = {
                s['key'] for s in _scan_bom_rows(df, header_row, col_part, col_qty, col_stock)
                if s['is_data']
            }
        if records_ok:
            _rebuild_bom_sheet_for_layer(wb[bom_sheet], df, groups)

    # ۴) گزارش لایه و قفل G4
    if records_ok:
        _add_layer_report_sheet(wb, layer, groups, bom_keys, placement_path)
    apply_g4_lock(wb, bom_sheet)

    wb.save(out_path)
    return out_path


def make_output_paths(bom_path: str, out_dir: str) -> dict[str, str]:
    """نام خروجی‌ها = نام فایل اصلی + v1 (با برچسب TOP/BOT برای نسخه‌های لایه‌ای)."""
    base = os.path.splitext(os.path.basename(bom_path))[0]
    return {
        "top": os.path.join(out_dir, f"{base}_TOP_v1.xlsx"),
        "bot": os.path.join(out_dir, f"{base}_BOT_v1.xlsx"),
        "bom": os.path.join(out_dir, f"{base}_v1.xlsx"),
    }


def build_all_outputs(bom_path: str, top_path: str, bot_path: str,
                      results: list[dict], out_dir: str,
                      top_sheet_name: str = "top",
                      bot_sheet_name: str = "bot") -> dict[str, str]:
    """
    ساخت هر سه خروجی:

    * <name>_TOP_v1.xlsx  — «BOM مختص لایهٔ TOP»: بر اساس کد انبار یکسان + تمام
      Designatorهای فایل TOP + تعداد واقعی آن‌ها (+ شیت top با مختصات PCB)
    * <name>_BOT_v1.xlsx  — همان برای لایهٔ BOT
    * <name>_v1.xlsx      — BOM بازتولیدشدهٔ کامل (دو لایه + گزارش اعتبارسنجی)

    در هر سه فایل، سلول G4 = «P.Parsa» و با رمز قفل است.
    """
    paths = make_output_paths(bom_path, out_dir)
    build_layer_workbook(bom_path, top_path, "top", paths["top"])
    build_layer_workbook(bom_path, bot_path, "bot", paths["bot"])
    build_combined_workbook(bom_path, top_path, bot_path, results, paths["bom"],
                            top_sheet_name=top_sheet_name,
                            bot_sheet_name=bot_sheet_name)
    return paths
